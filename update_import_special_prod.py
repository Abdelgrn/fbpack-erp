import os

# ==============================================================================
# 1. MISE À JOUR DU TEMPLATE HTML (Ajout option SPECIAL_PROD)
# ==============================================================================
import_html = """
{% extends 'base.html' %}
{% block content %}
<div class="max-w-3xl mx-auto mt-6">
    <h2 class="text-3xl font-bold text-white mb-2">Station d'Import Universelle</h2>
    <p class="text-gray-400 mb-6">Mettez à jour vos données en masse via Excel.</p>

    <div class="glass p-8 rounded-xl border border-slate-700">
        
        {% if message %}
        <div class="p-4 mb-6 text-sm rounded-lg font-bold border {% if success %}bg-green-900/50 border-green-500 text-green-300{% else %}bg-red-900/50 border-red-500 text-red-300{% endif %}">
            {{ message }}
        </div>
        {% endif %}

        {% if details %}
        <div class="mt-2 mb-6 p-4 bg-slate-900/50 border border-slate-700 rounded text-xs text-gray-300 max-h-60 overflow-y-auto">
            <p class="text-neon font-bold mb-2">📋 Détail ligne par ligne :</p>
            {% for d in details %}
                <p class="{% if 'ERREUR' in d %}text-red-400{% elif 'IGNORÉE' in d %}text-yellow-400{% else %}text-green-400{% endif %}">{{ d }}</p>
            {% endfor %}
        </div>
        {% endif %}

        <form method="post" enctype="multipart/form-data" class="space-y-6">
            {% csrf_token %}
            
            <!-- Choix du Type -->
            <div>
                <label class="block text-sm font-bold text-gray-300 mb-2">1. Que voulez-vous importer ?</label>
                <select name="import_type" id="typeSelector" onchange="updateInstructions()" class="w-full bg-slate-800 border border-slate-600 rounded p-3 text-white focus:border-neon focus:outline-none">
                    <option value="STOCK">📦 Stock (Matières Premières)</option>
                    <option value="CRM">🤝 CRM (Clients & Prospects)</option>
                    <option value="TOOLS">⚙️ Outillage (Cylindres & Clichés)</option>
                    <option value="PLANNING">🏭 Planning Production (OF)</option>
                    <option value="CONSO">💧 Consommation (Flexo/Hélio)</option>
                    <option value="SPECIAL_PROD">🔧 Production Spéciale (Découpe/Impression)</option>
                </select>
            </div>

            <!-- Zone Upload -->
            <div>
                <label class="block text-sm font-bold text-gray-300 mb-2">2. Fichier Excel (.xlsx)</label>
                <div class="border-2 border-dashed border-slate-600 rounded-lg p-8 text-center hover:border-neon transition bg-slate-800/30">
                    <input type="file" name="excel_file" accept=".xlsx, .xls" required
                           class="block w-full text-sm text-slate-500 file:mr-4 file:py-2 file:px-4 file:rounded-full file:border-0 file:text-sm file:font-semibold file:bg-neon file:text-black hover:file:bg-cyan-400"/>
                </div>
            </div>

            <!-- Instructions Dynamiques -->
            <div class="bg-slate-900/50 p-4 rounded text-sm border border-slate-700">
                <p class="text-neon font-bold mb-2">Colonnes Excel requises (En-têtes) :</p>
                <p id="colList" class="font-mono text-gray-300 text-xs leading-relaxed break-words">Select type...</p>
            </div>

            <!-- Exemple visuel pour SPECIAL_PROD -->
            <div id="exampleTable" class="hidden bg-slate-900/50 p-4 rounded text-sm border border-slate-700 overflow-x-auto">
                <p class="text-neon font-bold mb-2">📝 Exemple de votre fichier Excel :</p>
                <table class="text-xs text-gray-300 border-collapse w-full">
                    <thead>
                        <tr class="border-b border-slate-600">
                            <th class="p-1 text-left">Date</th>
                            <th class="p-1 text-left">Produit</th>
                            <th class="p-1 text-left">Support</th>
                            <th class="p-1 text-left">Qte_Lancee</th>
                            <th class="p-1 text-left">Lot</th>
                            <th class="p-1 text-left">Laize</th>
                            <th class="p-1 text-left">Client</th>
                            <th class="p-1 text-left">Equipe</th>
                            <th class="p-1 text-left">Machine</th>
                            <th class="p-1 text-left">H_Debut</th>
                            <th class="p-1 text-left">H_Fin</th>
                            <th class="p-1 text-left">Prod_ML</th>
                            <th class="p-1 text-left">Dec_Demarrage</th>
                            <th class="p-1 text-left">Dec_Lisiere</th>
                            <th class="p-1 text-left">Dec_Jonction</th>
                            <th class="p-1 text-left">Dec_Transport</th>
                            <th class="p-1 text-left">Prod_KG</th>
                            <th class="p-1 text-left">Rebobinage_KG</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr class="border-b border-slate-700">
                            <td class="p-1">15/01/2025</td>
                            <td class="p-1">Sac Lait</td>
                            <td class="p-1">PEBD 50μ</td>
                            <td class="p-1">500</td>
                            <td class="p-1">LOT-2025-001</td>
                            <td class="p-1">320</td>
                            <td class="p-1">Laiterie Atlas</td>
                            <td class="p-1">A</td>
                            <td class="p-1">IMP-01</td>
                            <td class="p-1">08:00</td>
                            <td class="p-1">16:30</td>
                            <td class="p-1">12000</td>
                            <td class="p-1">5.2</td>
                            <td class="p-1">3.1</td>
                            <td class="p-1">1.5</td>
                            <td class="p-1">2.0</td>
                            <td class="p-1">485</td>
                            <td class="p-1">10</td>
                        </tr>
                    </tbody>
                </table>
            </div>
            
            <button type="submit" class="w-full bg-gradient-to-r from-cyan-600 to-blue-600 text-white font-bold py-3 px-6 rounded shadow-lg hover:brightness-110 transition transform active:scale-95">
                Lancer l'Importation
            </button>
        </form>

        <!-- Bouton Télécharger Template -->
        <div class="mt-4">
            <a href="{% url 'download_template_special_prod' %}" 
               class="block text-center w-full bg-slate-700 text-neon font-bold py-3 px-6 rounded hover:bg-slate-600 transition">
                📥 Télécharger le Template Excel (Production Spéciale)
            </a>
        </div>
    </div>
</div>

<script>
    const colMap = {
        'STOCK': 'Designation, Categorie (Film/Encre/Colle), Quantite, Unite, Seuil_Min, Prix',
        'CRM': 'Nom, Ville, Telephone, Email, Secteur, Statut (Active/Prospect)',
        'TOOLS': 'Ref_Produit, Type (Cylindre/Cliche), Serial, Tours_Max, Tours_Actuels',
        'PLANNING': 'OF_Numero, Client, Produit, Machine, Date_Debut (JJ/MM/AAAA), Qte_Prevue',
        'CONSO': 'Date, Type (Flexo/Helio), Support, Laize, Bobine_In, Bobine_Out, Metrage, Noir, Magenta, Jaune, Cyan, Dore, Silver, Orange, Blanc, Vernis, Metoxyn, 2080',
        'SPECIAL_PROD': 'Date, Produit, Support, Qte_Lancee, Lot, Laize, Client, Equipe (A/B/C), Machine, H_Debut, H_Fin, Prod_ML, Dec_Demarrage, Dec_Lisiere, Dec_Jonction, Dec_Transport, Prod_KG, Rebobinage_KG'
    };

    function updateInstructions() {
        const type = document.getElementById('typeSelector').value;
        document.getElementById('colList').innerText = colMap[type];
        
        // Afficher/masquer l'exemple pour SPECIAL_PROD
        const exTable = document.getElementById('exampleTable');
        if (type === 'SPECIAL_PROD') {
            exTable.classList.remove('hidden');
        } else {
            exTable.classList.add('hidden');
        }
    }
    // Init
    updateInstructions();
</script>
{% endblock %}
"""

os.makedirs("templates/stock", exist_ok=True)
with open("templates/stock/import_stock.html", "w", encoding="utf-8") as f:
    f.write(import_html)
print("✅ Template HTML mis à jour avec option SPECIAL_PROD.")


# ==============================================================================
# 2. AJOUT DE LA VUE D'IMPORT SPECIAL_PROD + TEMPLATE DOWNLOAD dans views.py
# ==============================================================================
views_path = "core/views.py"

with open(views_path, "r", encoding="utf-8") as f:
    old_content = f.read()

# --- Le nouveau code complet de la vue import ---
new_view_code = '''
import pandas as pd
from django.core.files.storage import FileSystemStorage
from django.utils.dateparse import parse_datetime
import datetime
from django.http import HttpResponse
from .models import *

def parse_time_safe(val):
    """Parse une heure depuis Excel de manière robuste"""
    if val is None or val == 0 or val == '' or str(val).strip() == '':
        return None
    try:
        # Si c'est déjà un objet time
        if hasattr(val, 'hour'):
            return val
        # Si c'est un datetime pandas
        if hasattr(val, 'time'):
            return val.time()
        # Si c'est un float (Excel stocke les heures en fraction de jour)
        if isinstance(val, float):
            total_seconds = int(val * 24 * 3600)
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            return datetime.time(hours, minutes)
        # Si c'est un string "08:00" ou "8:00"
        s = str(val).strip()
        parts = s.replace('h', ':').replace('H', ':').split(':')
        if len(parts) >= 2:
            h = int(parts[0])
            m = int(parts[1])
            return datetime.time(h, m)
        return None
    except:
        return None

def import_stock_view(request):
    context = {}
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            import_type = request.POST.get('import_type')
            excel_file = request.FILES['excel_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)
            
            df = pd.read_excel(file_path).fillna(0)
            count = 0
            errors = 0
            details = []
            
            # --- 1. IMPORT STOCK ---
            if import_type == 'STOCK':
                for idx, row in df.iterrows():
                    try:
                        cat_map = {'Film': 'FILM', 'Encre': 'INK', 'Colle': 'GLUE', 'Solvant': 'SOLV'}
                        cat_val = row.get('Categorie', 'FILM')
                        if cat_val == 0: cat_val = 'FILM'
                        
                        Material.objects.update_or_create(
                            name=row.get('Designation', 'Inconnu'),
                            defaults={
                                'category': cat_map.get(cat_val, 'FILM'),
                                'quantity': row.get('Quantite', 0),
                                'unit': row.get('Unite', 'kg'),
                                'min_threshold': row.get('Seuil_Min', 0),
                                'price_per_unit': row.get('Prix', 0)
                            }
                        )
                        count += 1
                        details.append(f"Ligne {idx+2}: ✅ {row.get('Designation')} importé")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ ERREUR - {str(e)}")

            # --- 2. IMPORT CRM ---
            elif import_type == 'CRM':
                for idx, row in df.iterrows():
                    try:
                        status_map = {'Active': 'ACTIVE', 'Prospect': 'PROSPECT', 'VIP': 'VIP'}
                        nom = row.get('Nom')
                        if nom and nom != 0:
                            Client.objects.update_or_create(
                                name=nom,
                                defaults={
                                    'city': row.get('Ville', ''),
                                    'phone': row.get('Telephone', ''),
                                    'email': row.get('Email', ''),
                                    'sector': row.get('Secteur', ''),
                                    'status': status_map.get(row.get('Statut'), 'PROSPECT')
                                }
                            )
                            count += 1
                            details.append(f"Ligne {idx+2}: ✅ Client {nom} importé")
                        else:
                            details.append(f"Ligne {idx+2}: ⚠️ IGNORÉE - Nom vide")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ ERREUR - {str(e)}")

            # --- 3. IMPORT OUTILLAGE ---
            elif import_type == 'TOOLS':
                for idx, row in df.iterrows():
                    try:
                        prod_ref = row.get('Ref_Produit')
                        if prod_ref and prod_ref != 0:
                            cli, _ = Client.objects.get_or_create(name="Client Divers", defaults={'city': 'Interne', 'phone': '000'})
                            
                            product, _ = TechnicalProduct.objects.get_or_create(
                                ref_internal=prod_ref,
                                defaults={'name': f"Produit {prod_ref}", 'client': cli, 'structure_type': 'MONO', 'width_mm': 100}
                            )
                            
                            type_map = {'Cylindre': 'CYL', 'Cliche': 'CLICHE'}
                            type_val = row.get('Type', 'CYL')
                            if type_val == 0: type_val = 'CYL'

                            Tooling.objects.update_or_create(
                                serial_number=row.get('Serial'),
                                defaults={
                                    'product': product,
                                    'tool_type': type_map.get(type_val, 'CYL'),
                                    'max_impressions': row.get('Tours_Max', 1000000),
                                    'current_impressions': row.get('Tours_Actuels', 0)
                                }
                            )
                            count += 1
                            details.append(f"Ligne {idx+2}: ✅ Outil {row.get('Serial')} importé")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ ERREUR - {str(e)}")

            # --- 4. IMPORT PLANNING ---
            elif import_type == 'PLANNING':
                for idx, row in df.iterrows():
                    try:
                        cli_name = row.get('Client')
                        if cli_name and cli_name != 0:
                            client, _ = Client.objects.get_or_create(name=cli_name, defaults={'city': '?', 'phone': '?'})
                            
                            prod_name = row.get('Produit')
                            product, _ = TechnicalProduct.objects.get_or_create(
                                ref_internal=f"REF-{str(prod_name)[:5]}", 
                                defaults={'name': prod_name, 'client': client, 'structure_type': 'MONO', 'width_mm': 500}
                            )
                            
                            mac_name = row.get('Machine')
                            machine = Machine.objects.filter(name__icontains=str(mac_name)).first()
                            
                            try:
                                start_d = pd.to_datetime(row.get('Date_Debut'))
                            except:
                                start_d = datetime.datetime.now()
                            end_d = start_d + datetime.timedelta(hours=4)
                            
                            ProductionOrder.objects.update_or_create(
                                of_number=str(row.get('OF_Numero')),
                                defaults={
                                    'client': client, 'product': product, 'machine': machine,
                                    'start_time': start_d, 'end_time': end_d,
                                    'quantity_planned': row.get('Qte_Prevue', 0), 'status': 'PLANNED'
                                }
                            )
                            count += 1
                            details.append(f"Ligne {idx+2}: ✅ OF {row.get('OF_Numero')} importé")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ ERREUR - {str(e)}")

            # --- 5. IMPORT CONSOMMATION ---
            elif import_type == 'CONSO':
                for idx, row in df.iterrows():
                    try:
                        try:
                            d_prod = pd.to_datetime(row.get('Date')).date()
                        except:
                            d_prod = datetime.date.today()

                        raw_type = str(row.get('Type', 'FLEXO')).upper()
                        final_type = 'HELIO' if 'HELIO' in raw_type else 'FLEXO'

                        ConsommationEncre.objects.create(
                            date=d_prod, process_type=final_type,
                            support=row.get('Support', 'Inconnu'),
                            laize=float(row.get('Laize', 0)),
                            bobine_in=float(row.get('Bobine_In', 0)),
                            bobine_out=float(row.get('Bobine_Out', 0)),
                            metrage=float(row.get('Metrage', 0)),
                            encre_noir=float(row.get('Noir', 0)),
                            encre_magenta=float(row.get('Magenta', 0)),
                            encre_jaune=float(row.get('Jaune', 0)),
                            encre_cyan=float(row.get('Cyan', 0)),
                            encre_dore=float(row.get('Dore', 0)),
                            encre_silver=float(row.get('Silver', 0)),
                            encre_orange=float(row.get('Orange', 0)),
                            encre_blanc=float(row.get('Blanc', 0)),
                            encre_vernis=float(row.get('Vernis', 0)),
                            solvant_metoxyn=float(row.get('Metoxyn', 0)),
                            solvant_2080=float(row.get('2080', 0))
                        )
                        count += 1
                        details.append(f"Ligne {idx+2}: ✅ Conso {d_prod} importée")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ ERREUR - {str(e)}")

            # --- 6. IMPORT PRODUCTION SPÉCIALE (NOUVEAU) ---
            elif import_type == 'SPECIAL_PROD':
                for idx, row in df.iterrows():
                    try:
                        # A. Date
                        try:
                            d_val = pd.to_datetime(row.get('Date')).date()
                        except:
                            d_val = datetime.date.today()
                        
                        # B. Produit (obligatoire)
                        produit_val = str(row.get('Produit', '')).strip()
                        if not produit_val or produit_val == '0':
                            details.append(f"Ligne {idx+2}: ⚠️ IGNORÉE - Produit vide")
                            continue
                        
                        # C. Client (ForeignKey - on cherche ou on crée)
                        client_obj = None
                        cli_name = str(row.get('Client', '')).strip()
                        if cli_name and cli_name != '0' and cli_name != '':
                            client_obj, _ = Client.objects.get_or_create(
                                name=cli_name,
                                defaults={'city': 'Non renseigné', 'phone': '000'}
                            )
                        
                        # D. Machine (ForeignKey - on cherche par nom)
                        machine_obj = None
                        mac_name = str(row.get('Machine', '')).strip()
                        if mac_name and mac_name != '0' and mac_name != '':
                            machine_obj = Machine.objects.filter(name__icontains=mac_name).first()
                            if not machine_obj:
                                # Créer la machine si elle n'existe pas
                                machine_obj = Machine.objects.create(
                                    name=mac_name,
                                    type='IMP',
                                    status='STOP'
                                )
                                details.append(f"  → Machine '{mac_name}' créée automatiquement")
                        
                        # E. Équipe
                        equipe_val = str(row.get('Equipe', 'A')).strip().upper()
                        if equipe_val not in ['A', 'B', 'C']:
                            equipe_val = 'A'
                        
                        # F. Heures (gestion robuste)
                        h_debut = parse_time_safe(row.get('H_Debut'))
                        h_fin = parse_time_safe(row.get('H_Fin'))
                        
                        # G. Valeurs numériques (avec protection)
                        def safe_float(val, default=0):
                            try:
                                v = float(val)
                                return v if v == v else default  # NaN check
                            except:
                                return default
                        
                        # H. Création de l'entrée
                        ProductionEntry.objects.create(
                            date=d_val,
                            produit=produit_val,
                            support=str(row.get('Support', '')),
                            quantite_lancee=safe_float(row.get('Qte_Lancee')),
                            lot=str(row.get('Lot', '')),
                            laize=safe_float(row.get('Laize')),
                            client=client_obj,
                            equipe=equipe_val,
                            machine=machine_obj,
                            heure_debut=h_debut,
                            heure_fin=h_fin,
                            prod_ml=safe_float(row.get('Prod_ML')),
                            dechets_demarrage=safe_float(row.get('Dec_Demarrage')),
                            dechets_lisiere=safe_float(row.get('Dec_Lisiere')),
                            dechets_jonction=safe_float(row.get('Dec_Jonction')),
                            dechets_transport=safe_float(row.get('Dec_Transport')),
                            prod_kg=safe_float(row.get('Prod_KG')),
                            rebobinage_kg=safe_float(row.get('Rebobinage_KG'))
                        )
                        count += 1
                        details.append(f"Ligne {idx+2}: ✅ {produit_val} ({d_val}) importé")
                        
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ ERREUR - {str(e)}")

            # --- RÉSULTAT FINAL ---
            if errors > 0:
                context = {
                    'message': f'⚠️ Import terminé : {count} lignes OK, {errors} erreurs dans {import_type}.',
                    'success': count > 0,
                    'details': details
                }
            else:
                context = {
                    'message': f'✅ Succès total ! {count} lignes importées dans {import_type}.',
                    'success': True,
                    'details': details
                }
            
            # Nettoyage du fichier uploadé
            try:
                os.remove(file_path)
            except:
                pass
                
        except Exception as e:
            context = {'message': f'❌ Erreur critique : {str(e)}', 'success': False}
            
    return render(request, 'stock/import_stock.html', context)


def download_template_special_prod(request):
    """Génère et télécharge un fichier Excel template pré-rempli"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Speciale"
    
    # En-têtes
    headers = [
        'Date', 'Produit', 'Support', 'Qte_Lancee', 'Lot', 'Laize',
        'Client', 'Equipe', 'Machine', 'H_Debut', 'H_Fin', 'Prod_ML',
        'Dec_Demarrage', 'Dec_Lisiere', 'Dec_Jonction', 'Dec_Transport',
        'Prod_KG', 'Rebobinage_KG'
    ]
    
    # Style en-têtes
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Ligne d'exemple
    example = [
        '15/01/2025', 'Sac Lait 1L', 'PEBD 50μ', 500, 'LOT-2025-001', 320,
        'Laiterie Atlas', 'A', 'IMP-01', '08:00', '16:30', 12000,
        5.2, 3.1, 1.5, 2.0, 485, 10
    ]
    
    example_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = example_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Deuxième ligne d'exemple
    example2 = [
        '15/01/2025', 'Film Emballage', 'OPP 40μ', 300, 'LOT-2025-002', 450,
        'Conserverie Sud', 'B', 'IMP-02', '14:00', '22:00', 8500,
        3.0, 2.5, 1.0, 1.5, 290, 5
    ]
    
    example_fill2 = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    for col, val in enumerate(example2, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.fill = example_fill2
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Feuille d'instructions
    ws2 = wb.create_sheet("INSTRUCTIONS")
    instructions = [
        ["GUIDE D'UTILISATION - Import Production Spéciale"],
        [""],
        ["Colonne", "Description", "Format", "Obligatoire"],
        ["Date", "Date de production", "JJ/MM/AAAA", "✅ Oui"],
        ["Produit", "Nom du produit fabriqué", "Texte", "✅ Oui"],
        ["Support", "Type de support/matière", "Texte", "Non"],
        ["Qte_Lancee", "Quantité lancée en kg", "Nombre", "Non"],
        ["Lot", "Numéro de lot", "Texte", "Non"],
        ["Laize", "Laize en mm", "Nombre", "Non"],
        ["Client", "Nom du client (sera créé si inexistant)", "Texte", "Non"],
        ["Equipe", "Équipe de travail", "A, B ou C", "Non (défaut: A)"],
        ["Machine", "Nom de la machine (sera créée si inexistante)", "Texte", "Non"],
        ["H_Debut", "Heure de début", "HH:MM", "Non"],
        ["H_Fin", "Heure de fin", "HH:MM", "Non"],
        ["Prod_ML", "Production en mètres linéaires", "Nombre", "Non"],
        ["Dec_Demarrage", "Déchets démarrage en kg", "Nombre", "Non"],
        ["Dec_Lisiere", "Déchets lisière en kg", "Nombre", "Non"],
        ["Dec_Jonction", "Déchets jonction en kg", "Nombre", "Non"],
        ["Dec_Transport", "Déchets transport en kg", "Nombre", "Non"],
        ["Prod_KG", "Production totale en kg", "Nombre", "Non"],
        ["Rebobinage_KG", "Rebobinage en kg", "Nombre", "Non"],
        [""],
        ["⚠️ IMPORTANT:"],
        ["- La première ligne doit contenir les en-têtes EXACTEMENT comme indiqué"],
        ["- Supprimez les lignes d'exemple avant d'importer"],
        ["- Les clients et machines seront créés automatiquement s'ils n'existent pas"],
        ["- Les colonnes vides seront remplies avec des valeurs par défaut (0)"],
    ]
    
    title_font = Font(name='Arial', bold=True, size=14, color='0D47A1')
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = title_font
            if row_idx == 3:
                cell.font = Font(bold=True)
    
    # Ajuster les largeurs
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_sheet.column_dimensions[col[0].column_letter].width = max_len + 4
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Template_Production_Speciale.xlsx"'
    wb.save(response)
    return response
'''

# Insertion intelligente dans views.py
if "def import_stock_view" in old_content:
    # Trouver le début de la fonction
    idx = old_content.find("def import_stock_view")
    # Garder tout ce qui est AVANT cette fonction (imports, autres vues)
    before = old_content[:idx]
    
    # Supprimer les anciens imports pandas/FileSystemStorage s'ils existent déjà
    lines_before = before.split('\n')
    cleaned_lines = []
    skip_imports = ['import pandas', 'from django.core.files.storage import FileSystemStorage', 
                    'from django.utils.dateparse import parse_datetime', 'from django.http import HttpResponse']
    for line in lines_before:
        if not any(imp in line for imp in skip_imports):
            cleaned_lines.append(line)
    
    before_cleaned = '\n'.join(cleaned_lines)
    final_content = before_cleaned + new_view_code
else:
    final_content = old_content + "\n" + new_view_code

with open(views_path, "w", encoding="utf-8") as f:
    f.write(final_content)

print("✅ Views.py mis à jour avec import SPECIAL_PROD + download template.")


# ==============================================================================
# 3. MISE À JOUR DES URLS
# ==============================================================================
urls_path = "core/urls.py"

with open(urls_path, "r", encoding="utf-8") as f:
    urls_content = f.read()

# Ajouter l'URL du template download si elle n'existe pas
if "download_template_special_prod" not in urls_content:
    # Trouver le dernier ] de urlpatterns
    insert_before = urls_content.rfind(']')
    new_url = """    path('import/template-special-prod/', views.download_template_special_prod, name='download_template_special_prod'),\n"""
    urls_content = urls_content[:insert_before] + new_url + urls_content[insert_before:]
    
    with open(urls_path, "w", encoding="utf-8") as f:
        f.write(urls_content)
    print("✅ URL download_template_special_prod ajoutée.")
else:
    print("ℹ️ URL download_template_special_prod existe déjà.")


print("")
print("=" * 60)
print("🚀 TERMINÉ ! Voici ce qui a été ajouté :")
print("=" * 60)
print("1. ✅ Template HTML - Option 'Production Spéciale' dans le sélecteur")
print("2. ✅ Vue d'import - Gestion complète de ProductionEntry")
print("3. ✅ Gestion des ForeignKey (Client créé si inexistant)")
print("4. ✅ Gestion des ForeignKey (Machine créée si inexistante)")
print("5. ✅ Gestion robuste des heures (08:00, 8h00, float Excel)")
print("6. ✅ Template Excel téléchargeable avec exemples")
print("7. ✅ Rapport détaillé ligne par ligne après import")
print("8. ✅ URL ajoutée pour le download du template")
print("")
print("👉 Relancez le serveur: python manage.py runserver")