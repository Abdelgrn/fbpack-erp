import os
import datetime
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import timedelta
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Count, Q, Avg
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.admin.models import LogEntry
from django.db import models as django_models
from django.utils.html import format_html
import re

from .models import (
    Client, ClientContact, InteractionLog, Opportunite,
    TechnicalProduct, Tooling, Quote,
    ProductionOrder, Machine, ConsumptionLog,
    Material, Supplier, ConsommationEncre, ProductionEntry,
    PurchaseOrder,
    StockLocation, StockLot, StockMovement,
    DemandeAchat, BonCommande, LigneBonCommande, StockSeuil,
    OrdreFabrication, EtapeProduction, SemiProduit,
    SuiviProduction, ConsommationMatiere, ProcessType,
    ChatRoom, ChatMessage, UserPresence,
)
from .forms import (
    ClientForm, ClientContactForm, InteractionLogForm, OpportuniteForm,
    ProductForm, ToolForm, QuoteForm,
    ProductionOrderForm, SupplierForm, MaterialForm, ConsommationEncreForm,
    MachineForm, ProductionEntryForm,
    OrdreFabricationForm, EtapeProductionForm, EtapeProductionFormSet,
    SuiviProductionForm, SemiProduitForm, ConsommationMatiereForm,
    ProcessTypeForm, OFLancementRapideForm,
)


# ===========================================================================
# --- DASHBOARD ---
# ===========================================================================

@login_required
def dashboard(request):
    context = {
        'count_clients': Client.objects.count(),
        'count_of_running': ProductionOrder.objects.filter(status='IN_PROGRESS').count(),
        'low_stock_count': len([m for m in Material.objects.all() if m.is_low_stock()]),
        'machines': Machine.objects.all(),
        'orders_per_machine': ProductionOrder.objects.values('machine__name').annotate(count=Count('id')),
        'count_opportunites': Opportunite.objects.filter(status__in=['PROSPECT', 'QUALIFICATION', 'PROPOSITION', 'NEGOCIATION']).count(),
        'count_devis_envoyes': Quote.objects.filter(status='SENT').count(),
        'pipeline_total': Opportunite.objects.exclude(status__in=['GAGNE', 'PERDU']).aggregate(total=Sum('valeur_estimee'))['total'] or 0,
        # Nouvelles stats OF
        'of_total': OrdreFabrication.objects.count(),
        'of_en_cours': OrdreFabrication.objects.filter(statut='EN_COURS').count(),
        'of_en_retard': sum(1 for of in OrdreFabrication.objects.filter(
            statut__in=['LANCE', 'EN_COURS']
        ) if of.est_en_retard),
        'of_termine_mois': OrdreFabrication.objects.filter(
            statut='TERMINE',
            date_fin_reelle__month=timezone.now().month
        ).count(),
        'semi_produits_dispo': SemiProduit.objects.filter(statut='DISPONIBLE').count(),
        'of_recents': OrdreFabrication.objects.select_related(
            'client', 'produit'
        ).order_by('-date_creation')[:5],
    }
    return render(request, 'dashboard.html', context)


@login_required
def production_gantt(request):
    orders = ProductionOrder.objects.exclude(status='DONE').order_by('machine', 'start_time')
    return render(request, 'production/gantt.html', {'orders': orders})


@login_required
def reporting(request):
    delayed_ofs = ProductionOrder.objects.filter(status='LATE')
    top_clients = ProductionOrder.objects.values('client__name').annotate(total_kg=Sum('produced_qty')).order_by('-total_kg')[:5]
    top_consumptions = ConsumptionLog.objects.values('material__name').annotate(total_used=Sum('quantity_used')).order_by('-total_used')[:5]
    total_devis = Quote.objects.count()
    devis_acceptes = Quote.objects.filter(status__in=['ACCEPTED', 'SIGNED']).count()
    taux_conversion = round((devis_acceptes / total_devis * 100), 1) if total_devis > 0 else 0
    context = {
        'delayed_ofs': delayed_ofs,
        'top_clients_labels': [c['client__name'] for c in top_clients],
        'top_clients_data': [c['total_kg'] for c in top_clients],
        'top_consumptions_labels': [c['material__name'] for c in top_consumptions],
        'top_consumptions_data': [c['total_used'] for c in top_consumptions],
        'taux_conversion': taux_conversion,
        'total_devis': total_devis,
        'devis_acceptes': devis_acceptes,
    }
    return render(request, 'reporting.html', context)


# ===========================================================================
# --- CRM ---
# ===========================================================================

@login_required
def crm_view(request):
    status_filter = request.GET.get('status', '')
    segment_filter = request.GET.get('segment', '')
    region_filter = request.GET.get('region', '')
    search = request.GET.get('q', '')
    clients = Client.objects.all().order_by('name')
    if status_filter:
        clients = clients.filter(status=status_filter)
    if segment_filter:
        clients = clients.filter(segment=segment_filter)
    if region_filter:
        clients = clients.filter(region=region_filter)
    if search:
        clients = clients.filter(
            Q(name__icontains=search) |
            Q(city__icontains=search) |
            Q(code_client__icontains=search)
        )
    quotes = Quote.objects.all().order_by('-date')
    opportunites = Opportunite.objects.all().order_by('-date_ouverture')
    pipeline_stages = []
    for stage_code, stage_label in Opportunite.STAGE_CHOICES:
        count = Opportunite.objects.filter(status=stage_code).count()
        montant = Opportunite.objects.filter(status=stage_code).aggregate(t=Sum('valeur_estimee'))['t'] or 0
        pipeline_stages.append({
            'code': stage_code,
            'label': stage_label,
            'count': count,
            'montant': montant
        })
    context = {
        'clients': clients,
        'quotes': quotes,
        'opportunites': opportunites,
        'pipeline_stages': pipeline_stages,
        'status_filter': status_filter,
        'segment_filter': segment_filter,
        'region_filter': region_filter,
        'search': search,
        'status_choices': Client.STATUS_CHOICES,
        'segment_choices': Client.SEGMENT_CHOICES,
        'region_choices': Client.REGION_CHOICES,
    }
    return render(request, 'crm.html', context)


@login_required
def client_detail(request, id):
    client = get_object_or_404(Client, id=id)
    contacts = ClientContact.objects.filter(client=client)
    interactions = InteractionLog.objects.filter(client=client).order_by('-date')[:20]
    opportunites = Opportunite.objects.filter(client=client).order_by('-date_ouverture')
    quotes = Quote.objects.filter(client=client).order_by('-date')
    orders = ProductionOrder.objects.filter(client=client).order_by('-start_time')[:5]
    ofs = OrdreFabrication.objects.filter(client=client).order_by('-date_creation')[:5]
    context = {
        'client': client,
        'contacts': contacts,
        'interactions': interactions,
        'opportunites': opportunites,
        'quotes': quotes,
        'orders': orders,
        'ofs': ofs,
    }
    return render(request, 'crm/client_detail.html', context)


@login_required
def add_client(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save()
            messages.success(request, f"Client « {client.name} » créé.")
            return redirect('client_detail', id=client.id)
    else:
        form = ClientForm()
    return render(request, 'crm/client_form.html', {'form': form, 'titre': 'Nouveau Client'})


@login_required
def edit_client(request, id):
    client = get_object_or_404(Client, id=id)
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, "Client mis à jour.")
            return redirect('client_detail', id=client.id)
    else:
        form = ClientForm(instance=client)
    return render(request, 'crm/client_form.html', {
        'form': form,
        'titre': f'Modifier {client.name}',
        'client': client
    })


@login_required
def add_contact(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if request.method == 'POST':
        form = ClientContactForm(request.POST)
        if form.is_valid():
            contact = form.save(commit=False)
            contact.client = client
            contact.save()
            messages.success(request, f"Contact « {contact.name} » ajouté.")
            return redirect('client_detail', id=client_id)
    else:
        form = ClientContactForm()
    return render(request, 'crm/contact_form.html', {
        'form': form,
        'client': client,
        'titre': 'Nouveau Contact'
    })


@login_required
def edit_contact(request, id):
    contact = get_object_or_404(ClientContact, id=id)
    if request.method == 'POST':
        form = ClientContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()
            messages.success(request, "Contact mis à jour.")
            return redirect('client_detail', id=contact.client.id)
    else:
        form = ClientContactForm(instance=contact)
    return render(request, 'crm/contact_form.html', {
        'form': form,
        'client': contact.client,
        'titre': f'Modifier {contact.name}'
    })


@login_required
def delete_contact(request, id):
    contact = get_object_or_404(ClientContact, id=id)
    client_id = contact.client.id
    if request.method == 'POST':
        contact.delete()
        messages.success(request, "Contact supprimé.")
    return redirect('client_detail', id=client_id)


@login_required
def add_interaction(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if request.method == 'POST':
        form = InteractionLogForm(request.POST, client=client)
        if form.is_valid():
            interaction = form.save(commit=False)
            interaction.client = client
            interaction.save()
            messages.success(request, "Interaction enregistrée.")
            return redirect('client_detail', id=client_id)
    else:
        form = InteractionLogForm(client=client)
    return render(request, 'crm/interaction_form.html', {
        'form': form,
        'client': client,
        'titre': 'Nouvelle Interaction'
    })


@login_required
def opportunites_view(request):
    opportunites = Opportunite.objects.all().order_by('-date_ouverture')
    pipeline = {}
    for stage_code, stage_label in Opportunite.STAGE_CHOICES:
        pipeline[stage_code] = {
            'label': stage_label,
            'items': Opportunite.objects.filter(status=stage_code),
            'total': Opportunite.objects.filter(status=stage_code).aggregate(
                t=Sum('valeur_estimee'))['t'] or 0,
        }
    return render(request, 'crm/opportunites.html', {
        'opportunites': opportunites,
        'pipeline': pipeline
    })


@login_required
def add_opportunite(request):
    if request.method == 'POST':
        form = OpportuniteForm(request.POST)
        if form.is_valid():
            opp = form.save()
            messages.success(request, f"Opportunité « {opp.titre} » créée.")
            return redirect('crm_view')
    else:
        initial = {}
        client_id = request.GET.get('client_id')
        if client_id:
            initial['client'] = client_id
        form = OpportuniteForm(initial=initial)
    return render(request, 'crm/opportunite_form.html', {
        'form': form,
        'titre': 'Nouvelle Opportunité'
    })


@login_required
def edit_opportunite(request, id):
    opp = get_object_or_404(Opportunite, id=id)
    if request.method == 'POST':
        form = OpportuniteForm(request.POST, instance=opp)
        if form.is_valid():
            form.save()
            messages.success(request, "Opportunité mise à jour.")
            return redirect('crm_view')
    else:
        form = OpportuniteForm(instance=opp)
    return render(request, 'crm/opportunite_form.html', {
        'form': form,
        'opp': opp,
        'titre': f'Modifier : {opp.titre}'
    })


@login_required
def delete_opportunite(request, id):
    opp = get_object_or_404(Opportunite, id=id)
    if request.method == 'POST':
        opp.delete()
        messages.success(request, "Opportunité supprimée.")
    return redirect('crm_view')


@login_required
def quotes_view(request):
    quotes = Quote.objects.all().order_by('-date')
    return render(request, 'crm/quotes_list.html', {'quotes': quotes})


@login_required
def add_quote(request):
    if request.method == 'POST':
        form = QuoteForm(request.POST, request.FILES)
        if form.is_valid():
            quote = form.save()
            messages.success(request, f"Devis {quote.reference} créé.")
            return redirect('crm_view')
    else:
        initial = {}
        client_id = request.GET.get('client_id')
        if client_id:
            initial['client'] = client_id
        form = QuoteForm(initial=initial)
    return render(request, 'crm/quote_form.html', {'form': form, 'titre': 'Nouveau Devis'})


@login_required
def edit_quote(request, id):
    quote = get_object_or_404(Quote, id=id)
    if request.method == 'POST':
        form = QuoteForm(request.POST, request.FILES, instance=quote)
        if form.is_valid():
            form.save()
            messages.success(request, "Devis mis à jour.")
            return redirect('crm_view')
    else:
        form = QuoteForm(instance=quote)
    return render(request, 'crm/quote_form.html', {
        'form': form,
        'titre': f'Modifier Devis {quote.reference}'
    })


@login_required
def convert_quote_to_order(request, id):
    quote = get_object_or_404(Quote, id=id)
    if quote.status in ['ACCEPTED', 'SIGNED'] and not quote.commande_creee:
        quote.commande_creee = True
        quote.save()
        messages.success(request, f"Devis {quote.reference} converti en commande.")
    return redirect('crm_view')


# ===========================================================================
# --- PRÉPRESSE ---
# ===========================================================================

@login_required
def prepress_view(request):
    products = TechnicalProduct.objects.all().order_by('-id')
    tools = Tooling.objects.all().order_by('-id')
    return render(request, 'prepress.html', {'products': products, 'tools': tools})


@login_required
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('prepress_view')
    else:
        form = ProductForm()
    return render(request, 'product_form.html', {'form': form, 'titre': 'Nouveau Produit Technique'})


@login_required
def edit_product(request, id):
    product = get_object_or_404(TechnicalProduct, id=id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('prepress_view')
    else:
        form = ProductForm(instance=product)
    return render(request, 'product_form.html', {
        'form': form,
        'titre': f'Modifier {product.ref_internal}'
    })


@login_required
def add_tool(request):
    if request.method == 'POST':
        form = ToolForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('prepress_view')
    else:
        form = ToolForm()
    return render(request, 'tool_form.html', {'form': form, 'titre': 'Nouvel Outillage'})


@login_required
def edit_tool(request, id):
    tool = get_object_or_404(Tooling, id=id)
    if request.method == 'POST':
        form = ToolForm(request.POST, instance=tool)
        if form.is_valid():
            form.save()
            return redirect('prepress_view')
    else:
        form = ToolForm(instance=tool)
    return render(request, 'tool_form.html', {
        'form': form,
        'titre': f'Modifier Outillage {tool.serial_number}'
    })


# ===========================================================================
# --- PRODUCTION (OF ANCIEN) ---
# ===========================================================================

@login_required
def production_view(request):
    ofs = ProductionOrder.objects.all().order_by('-start_time')
    return render(request, 'production_list.html', {'ofs': ofs})


@login_required
def add_production(request):
    if request.method == 'POST':
        form = ProductionOrderForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('production_view')
    else:
        form = ProductionOrderForm()
    return render(request, 'production_form.html', {
        'form': form,
        'titre': 'Créer un Ordre de Fabrication'
    })


@login_required
def edit_production(request, id):
    of = get_object_or_404(ProductionOrder, id=id)
    if request.method == 'POST':
        form = ProductionOrderForm(request.POST, request.FILES, instance=of)
        if form.is_valid():
            form.save()
            return redirect('production_view')
    else:
        form = ProductionOrderForm(instance=of)
    return render(request, 'production_form.html', {
        'form': form,
        'titre': f'Modifier OF {of.of_number}'
    })


# ===========================================================================
# --- OF MULTI-PROCESSUS (NOUVEAU) ---
# ===========================================================================

@login_required
def of_list_view(request):
    """Liste des Ordres de Fabrication avec filtres"""
    
    statut = request.GET.get('statut', '')
    priorite = request.GET.get('priorite', '')
    client_id = request.GET.get('client', '')
    search = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    ofs = OrdreFabrication.objects.select_related(
        'client', 'produit', 'cree_par'
    ).prefetch_related('etapes').order_by('-date_creation')
    
    if statut:
        ofs = ofs.filter(statut=statut)
    if priorite:
        ofs = ofs.filter(priorite=priorite)
    if client_id:
        ofs = ofs.filter(client_id=client_id)
    if search:
        ofs = ofs.filter(
            Q(numero_of__icontains=search) |
            Q(numero_lot__icontains=search) |
            Q(produit__name__icontains=search) |
            Q(client__name__icontains=search)
        )
    if date_from:
        ofs = ofs.filter(date_lancement__gte=date_from)
    if date_to:
        ofs = ofs.filter(date_lancement__lte=date_to)
    
    stats = {
        'total': ofs.count(),
        'brouillon': ofs.filter(statut='BROUILLON').count(),
        'en_cours': ofs.filter(statut='EN_COURS').count(),
        'termine': ofs.filter(statut='TERMINE').count(),
        'en_retard': sum(1 for of in ofs if of.est_en_retard),
    }
    
    context = {
        'ofs': ofs[:100],
        'stats': stats,
        'statut_choices': OrdreFabrication.STATUT_CHOICES,
        'priorite_choices': OrdreFabrication.PRIORITE_CHOICES,
        'clients': Client.objects.filter(status='ACTIVE'),
        'selected_statut': statut,
        'selected_priorite': priorite,
        'selected_client': client_id,
        'search': search,
        'selected_date_from': date_from,
        'selected_date_to': date_to,
    }
    return render(request, 'of/of_list.html', context)


@login_required
def of_create_view(request):
    """Créer un nouvel OF avec ses étapes"""
    
    if request.method == 'POST':
        form = OrdreFabricationForm(request.POST, request.FILES)
        formset = EtapeProductionFormSet(request.POST, prefix='etapes')
        
        if form.is_valid():
            of = form.save(commit=False)
            of.cree_par = request.user
            of.save()
            
            if formset.is_valid():
                etapes = formset.save(commit=False)
                for etape in etapes:
                    etape.of = of
                    etape.save()
                for obj in formset.deleted_objects:
                    obj.delete()
            
            messages.success(request, f"OF {of.numero_of} créé avec succès !")
            return redirect('of_detail', of_id=of.id)
        else:
            messages.error(request, "Erreur dans le formulaire. Vérifiez les champs.")
    else:
        form = OrdreFabricationForm()
        formset = EtapeProductionFormSet(prefix='etapes', queryset=EtapeProduction.objects.none())
    
    context = {
        'form': form,
        'formset': formset,
        'process_types': ProcessType.objects.filter(est_actif=True),
        'machines': Machine.objects.all(),
        'titre': 'Nouvel Ordre de Fabrication',
    }
    return render(request, 'of/of_form.html', context)


@login_required
def of_detail_view(request, of_id):
    """Détail d'un OF avec toutes ses étapes et suivi"""
    
    of = get_object_or_404(
        OrdreFabrication.objects.select_related('client', 'produit', 'cree_par'),
        id=of_id
    )
    
    etapes = of.etapes.select_related(
        'process_type', 'machine', 'operateur'
    ).prefetch_related('suivis', 'consommations').order_by('numero_etape')
    
    semi_produits = SemiProduit.objects.filter(of_origine=of).order_by('-date_creation')
    
    suivis = SuiviProduction.objects.filter(
        etape__of=of
    ).select_related('etape', 'operateur').order_by('-date_heure')[:50]
    
    etapes_data = []
    for etape in etapes:
        etapes_data.append({
            'nom': etape.get_nom_display(),
            'progression': etape.progression,
            'statut': etape.statut,
            'couleur': etape.get_statut_color(),
        })
    
    context = {
        'of': of,
        'etapes': etapes,
        'semi_produits': semi_produits,
        'suivis': suivis,
        'etapes_data_json': json.dumps(etapes_data),
    }
    return render(request, 'of/of_detail.html', context)


@login_required
def of_edit_view(request, of_id):
    """Modifier un OF existant"""
    
    of = get_object_or_404(OrdreFabrication, id=of_id)
    
    if request.method == 'POST':
        form = OrdreFabricationForm(request.POST, request.FILES, instance=of)
        formset = EtapeProductionFormSet(request.POST, instance=of, prefix='etapes')
        
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, f"OF {of.numero_of} mis à jour !")
            return redirect('of_detail', of_id=of.id)
        else:
            messages.error(request, "Erreur dans le formulaire.")
    else:
        form = OrdreFabricationForm(instance=of)
        formset = EtapeProductionFormSet(instance=of, prefix='etapes')
    
    context = {
        'form': form,
        'formset': formset,
        'of': of,
        'process_types': ProcessType.objects.filter(est_actif=True),
        'machines': Machine.objects.all(),
        'titre': f'Modifier OF {of.numero_of}',
    }
    return render(request, 'of/of_form.html', context)


@login_required
def of_delete_view(request, of_id):
    """Supprimer un OF"""
    
    of = get_object_or_404(OrdreFabrication, id=of_id)
    
    if request.method == 'POST':
        numero = of.numero_of
        of.delete()
        messages.success(request, f"OF {numero} supprimé.")
        return redirect('of_list')
    
    return render(request, 'of/of_confirm_delete.html', {'of': of})


@login_required
def of_changer_statut(request, of_id, nouveau_statut):
    """Changer le statut d'un OF"""
    
    of = get_object_or_404(OrdreFabrication, id=of_id)
    ancien_statut = of.statut
    
    if nouveau_statut in dict(OrdreFabrication.STATUT_CHOICES):
        of.statut = nouveau_statut
        
        if nouveau_statut == 'LANCE' and not of.date_lancement:
            of.date_lancement = timezone.now().date()
        elif nouveau_statut == 'TERMINE':
            of.date_fin_reelle = timezone.now().date()
        
        of.save()
        messages.success(
            request, 
            f"OF {of.numero_of} : {ancien_statut} → {nouveau_statut}"
        )
    else:
        messages.error(request, "Statut invalide.")
    
    return redirect('of_detail', of_id=of_id)


@login_required
def of_lancement_rapide(request):
    """Formulaire de lancement rapide d'OF avec étapes prédéfinies"""
    
    if request.method == 'POST':
        form = OFLancementRapideForm(request.POST)
        
        if form.is_valid():
            of = OrdreFabrication.objects.create(
                client=form.cleaned_data['client'],
                produit=form.cleaned_data['produit'],
                quantite_prevue=form.cleaned_data['quantite'],
                priorite=form.cleaned_data['priorite'],
                date_lancement=form.cleaned_data['date_lancement'],
                statut='LANCE',
                cree_par=request.user,
            )
            
            numero = 1
            
            if form.cleaned_data.get('etape_extrusion'):
                process = ProcessType.objects.filter(code='EXTRUSION').first()
                EtapeProduction.objects.create(
                    of=of,
                    numero_etape=numero,
                    process_type=process,
                    machine=form.cleaned_data.get('machine_extrusion'),
                    quantite_entree=form.cleaned_data.get('qte_extrusion') or form.cleaned_data['quantite'],
                    statut='PRET',
                )
                numero += 1
            
            if form.cleaned_data.get('etape_impression'):
                process = ProcessType.objects.filter(code='IMPRESSION').first()
                EtapeProduction.objects.create(
                    of=of,
                    numero_etape=numero,
                    process_type=process,
                    machine=form.cleaned_data.get('machine_impression'),
                    quantite_entree=form.cleaned_data.get('qte_impression') or form.cleaned_data['quantite'],
                    statut='EN_ATTENTE',
                )
                numero += 1
            
            if form.cleaned_data.get('etape_decoupe'):
                process = ProcessType.objects.filter(code='DECOUPE').first()
                EtapeProduction.objects.create(
                    of=of,
                    numero_etape=numero,
                    process_type=process,
                    machine=form.cleaned_data.get('machine_decoupe'),
                    quantite_entree=form.cleaned_data.get('qte_decoupe') or form.cleaned_data['quantite'],
                    statut='EN_ATTENTE',
                )
            
            messages.success(request, f"OF {of.numero_of} lancé avec {of.nb_etapes} étapes !")
            return redirect('of_detail', of_id=of.id)
    else:
        form = OFLancementRapideForm()
    
    context = {
        'form': form,
        'titre': 'Lancement Rapide OF',
    }
    return render(request, 'of/of_lancement_rapide.html', context)


# ===========================================================================
# --- ÉTAPES DE PRODUCTION ---
# ===========================================================================

@login_required
def etape_detail_view(request, etape_id):
    """Détail d'une étape avec son suivi"""
    
    etape = get_object_or_404(
        EtapeProduction.objects.select_related('of', 'process_type', 'machine', 'operateur'),
        id=etape_id
    )
    
    suivis = etape.suivis.select_related('operateur').order_by('-date_heure')
    consommations = etape.consommations.select_related('material', 'lot')
    semi_produits = SemiProduit.objects.filter(etape_origine=etape)
    
    if request.method == 'POST':
        form = SuiviProductionForm(request.POST)
        if form.is_valid():
            suivi = form.save(commit=False)
            suivi.etape = etape
            suivi.operateur = request.user
            suivi.save()
            
            etape.quantite_sortie += suivi.quantite_produite
            etape.quantite_rebut += suivi.quantite_rebut
            
            if suivi.type_evenement == 'DEMARRAGE':
                etape.statut = 'EN_COURS'
                if not etape.date_debut_reel:
                    etape.date_debut_reel = timezone.now()
            elif suivi.type_evenement == 'FIN':
                etape.statut = 'TERMINE'
                etape.date_fin_reel = timezone.now()
            elif suivi.type_evenement == 'ARRET':
                etape.statut = 'PAUSE'
            
            etape.save()
            
            of = etape.of
            of.quantite_produite = sum(e.quantite_sortie for e in of.etapes.filter(statut='TERMINE'))
            of.quantite_rebut = sum(e.quantite_rebut for e in of.etapes.all())
            of.save()
            
            messages.success(request, "Suivi enregistré !")
            return redirect('etape_detail', etape_id=etape_id)
    else:
        form = SuiviProductionForm()
    
    context = {
        'etape': etape,
        'of': etape.of,
        'suivis': suivis,
        'consommations': consommations,
        'semi_produits': semi_produits,
        'form': form,
    }
    return render(request, 'of/etape_detail.html', context)


@login_required
def etape_demarrer(request, etape_id):
    """Démarrer une étape"""
    
    etape = get_object_or_404(EtapeProduction, id=etape_id)
    
    if etape.statut in ['EN_ATTENTE', 'PRET', 'PAUSE']:
        etape.statut = 'EN_COURS'
        if not etape.date_debut_reel:
            etape.date_debut_reel = timezone.now()
        etape.save()
        
        SuiviProduction.objects.create(
            etape=etape,
            operateur=request.user,
            type_evenement='DEMARRAGE',
            commentaire="Étape démarrée"
        )
        
        if etape.of.statut == 'LANCE':
            etape.of.statut = 'EN_COURS'
            etape.of.save()
        
        messages.success(request, f"Étape {etape.numero_etape} démarrée !")
    
    return redirect('etape_detail', etape_id=etape_id)


@login_required
def etape_terminer(request, etape_id):
    """Terminer une étape et créer le semi-produit"""
    
    etape = get_object_or_404(EtapeProduction, id=etape_id)
    
    if request.method == 'POST':
        quantite_sortie = float(request.POST.get('quantite_sortie', 0))
        quantite_rebut = float(request.POST.get('quantite_rebut', 0))
        
        etape.quantite_sortie = quantite_sortie
        etape.quantite_rebut = quantite_rebut
        etape.statut = 'TERMINE'
        etape.date_fin_reel = timezone.now()
        etape.save()
        
        if etape.genere_semi_produit and quantite_sortie > 0:
            type_sp = 'FILM_EXTRUDE'
            if etape.process_type:
                if 'IMP' in etape.process_type.code.upper():
                    type_sp = 'FILM_IMPRIME'
                elif 'COMP' in etape.process_type.code.upper():
                    type_sp = 'FILM_COMPLEXE'
            
            SemiProduit.objects.create(
                designation=f"SP - {etape.of.produit.name} - Étape {etape.numero_etape}",
                type_semi_produit=type_sp,
                of_origine=etape.of,
                etape_origine=etape,
                quantite=quantite_sortie,
                laize=etape.of.laize,
                conforme=True,
            )
        
        SuiviProduction.objects.create(
            etape=etape,
            operateur=request.user,
            type_evenement='FIN',
            quantite_produite=quantite_sortie,
            quantite_rebut=quantite_rebut,
            commentaire="Étape terminée"
        )
        
        of = etape.of
        if all(e.statut == 'TERMINE' for e in of.etapes.all()):
            of.statut = 'TERMINE'
            of.date_fin_reelle = timezone.now().date()
            of.quantite_produite = quantite_sortie
        
        of.quantite_rebut = sum(e.quantite_rebut for e in of.etapes.all())
        of.save()
        
        etape_suivante = EtapeProduction.objects.filter(
            of=of,
            numero_etape=etape.numero_etape + 1
        ).first()
        
        if etape_suivante:
            etape_suivante.statut = 'PRET'
            etape_suivante.quantite_entree = quantite_sortie
            etape_suivante.save()
        
        messages.success(request, f"Étape {etape.numero_etape} terminée !")
        return redirect('of_detail', of_id=etape.of.id)
    
    return render(request, 'of/etape_terminer.html', {'etape': etape})


# ===========================================================================
# --- SEMI-PRODUITS ---
# ===========================================================================

@login_required
def semi_produit_list(request):
    """Liste des semi-produits"""
    
    statut = request.GET.get('statut', '')
    type_sp = request.GET.get('type', '')
    
    semi_produits = SemiProduit.objects.select_related(
        'of_origine', 'etape_origine', 'emplacement'
    ).order_by('-date_creation')
    
    if statut:
        semi_produits = semi_produits.filter(statut=statut)
    if type_sp:
        semi_produits = semi_produits.filter(type_semi_produit=type_sp)
    
    stats = {
        'total': semi_produits.count(),
        'disponible': semi_produits.filter(statut='DISPONIBLE').count(),
        'reserve': semi_produits.filter(statut='RESERVE').count(),
        'total_kg': semi_produits.filter(statut='DISPONIBLE').aggregate(
            t=Sum('quantite')
        )['t'] or 0,
    }
    
    context = {
        'semi_produits': semi_produits[:100],
        'stats': stats,
        'statut_choices': SemiProduit.STATUT_CHOICES,
        'type_choices': SemiProduit.TYPE_CHOICES,
        'selected_statut': statut,
        'selected_type': type_sp,
    }
    return render(request, 'of/semi_produit_list.html', context)


@login_required
def semi_produit_detail(request, sp_id):
    """Détail d'un semi-produit"""
    
    sp = get_object_or_404(
        SemiProduit.objects.select_related(
            'of_origine', 'etape_origine', 'etape_destination', 'emplacement'
        ),
        id=sp_id
    )
    
    context = {'semi_produit': sp}
    return render(request, 'of/semi_produit_detail.html', context)


# ===========================================================================
# --- PROCESS TYPES ---
# ===========================================================================

@login_required
def process_type_list(request):
    """Liste et gestion des types de processus"""
    
    process_types = ProcessType.objects.all().order_by('ordre_defaut')
    
    if request.method == 'POST':
        form = ProcessTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Type de processus créé !")
            return redirect('process_type_list')
    else:
        form = ProcessTypeForm()
    
    context = {
        'process_types': process_types,
        'form': form,
    }
    return render(request, 'of/process_type_list.html', context)


@login_required
def process_type_delete(request, pt_id):
    """Supprimer un type de processus"""
    
    pt = get_object_or_404(ProcessType, id=pt_id)
    if request.method == 'POST':
        pt.delete()
        messages.success(request, "Type de processus supprimé.")
    return redirect('process_type_list')


# ===========================================================================
# --- API JSON ---
# ===========================================================================

@login_required
def of_stats_api(request):
    """API JSON pour statistiques OF"""
    
    from collections import defaultdict
    
    statuts = {}
    for code, label in OrdreFabrication.STATUT_CHOICES:
        statuts[code] = OrdreFabrication.objects.filter(statut=code).count()
    
    date_30j = timezone.now().date() - timedelta(days=30)
    ofs_recents = OrdreFabrication.objects.filter(
        date_lancement__gte=date_30j
    ).values('date_lancement').annotate(
        qte=Sum('quantite_produite')
    ).order_by('date_lancement')
    
    prod_par_jour = {str(of['date_lancement']): of['qte'] or 0 for of in ofs_recents}
    
    top_clients = OrdreFabrication.objects.values(
        'client__name'
    ).annotate(
        total=Sum('quantite_prevue')
    ).order_by('-total')[:5]
    
    data = {
        'statuts': statuts,
        'production_par_jour': prod_par_jour,
        'top_clients': list(top_clients),
    }
    
    return JsonResponse(data)


# ===========================================================================
# --- STOCKS ---
# ===========================================================================

@login_required
def stock_view(request):
    materials = Material.objects.all()
    suppliers = Supplier.objects.all()
    consos = ConsommationEncre.objects.all().order_by('-date')
    return render(request, 'stock_list.html', {
        'materials': materials,
        'suppliers': suppliers,
        'consos': consos
    })


@login_required
def add_material(request):
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock_view')
    else:
        form = MaterialForm()
    return render(request, 'stock_list.html', {'form': form, 'titre': 'Nouveau Matériau'})


@login_required
def add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock_view')
    else:
        form = SupplierForm()
    return render(request, 'stock_list.html', {'form': form, 'titre': 'Nouveau Fournisseur'})


@login_required
def add_consommation(request):
    if request.method == 'POST':
        form = ConsommationEncreForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('conso_list')
    else:
        form = ConsommationEncreForm()
    return render(request, 'stock_list.html', {'form': form, 'titre': 'Nouvelle Consommation'})


@login_required
def conso_list_view(request):
    consos = ConsommationEncre.objects.all().order_by('-date')
    return render(request, 'stock_list.html', {'consos': consos})


# ===========================================================================
# --- STOCK AVANCÉ AVEC RECHERCHE INTELLIGENTE ---
# ===========================================================================

def highlight_search(text, query):
    """Surligne les termes recherchés dans le HTML"""
    if not query:
        return text
    pattern = re.compile(f'({re.escape(query)})', re.IGNORECASE)
    return pattern.sub(r'<mark class="bg-yellow-400 text-black px-1 rounded">\1</mark>', text)


@login_required
def stock_advanced_view(request):
    """Vue principale du module stock avancé avec RECHERCHE INTELLIGENTE"""
    
    # ══════════════════════════════════════════════════════════════
    # RÉCUPÉRATION DES FILTRES DE RECHERCHE
    # ══════════════════════════════════════════════════════════════
    search_query = request.GET.get('q', '').strip()
    category_filter = request.GET.get('category', '')
    low_stock_only = request.GET.get('low_stock', '') == 'on'
    supplier_filter = request.GET.get('supplier', '')
    
    # ══════════════════════════════════════════════════════════════
    # MATIÈRES PREMIÈRES AVEC FILTRES
    # ══════════════════════════════════════════════════════════════
    materials = Material.objects.select_related('supplier').all()
    
    # Appliquer les filtres de recherche
    if search_query:
        materials = materials.filter(
            Q(name__icontains=search_query) |
            Q(supplier__name__icontains=search_query)
        )
    
    if category_filter:
        materials = materials.filter(category=category_filter)
    
    if supplier_filter:
        materials = materials.filter(supplier_id=supplier_filter)
    
    materials = materials.order_by('name')
    
    # Filtrer par stock faible (après le queryset principal)
    if low_stock_only:
        materials = [m for m in materials if m.is_low_stock()]
    
    # ══════════════════════════════════════════════════════════════
    # CALCUL DES ALERTES STOCK
    # ══════════════════════════════════════════════════════════════
    all_materials = Material.objects.select_related('supplier').all()
    alertes_stock = []
    nb_ruptures = 0
    nb_critiques = 0
    nb_alertes_simples = 0
    
    # Compteurs par catégorie
    cat_stats = {
        'FILM': {'rupture': 0, 'critique': 0, 'alerte': 0},
        'INK': {'rupture': 0, 'critique': 0, 'alerte': 0},
        'GLUE': {'rupture': 0, 'critique': 0, 'alerte': 0},
        'SOLV': {'rupture': 0, 'critique': 0, 'alerte': 0},
    }
    
    for m in all_materials:
        if m.is_low_stock():
            # Calcul du pourcentage et niveau de criticité
            if m.min_threshold > 0:
                pct = round((m.quantity / m.min_threshold) * 100, 1)
            else:
                pct = 0
            
            if m.quantity <= 0:
                niveau = 'RUPTURE'
                icone = '🔴'
                nb_ruptures += 1
                if m.category in cat_stats:
                    cat_stats[m.category]['rupture'] += 1
            elif pct < 50:
                niveau = 'CRITIQUE'
                icone = '🟠'
                nb_critiques += 1
                if m.category in cat_stats:
                    cat_stats[m.category]['critique'] += 1
            else:
                niveau = 'ALERTE'
                icone = '🟡'
                nb_alertes_simples += 1
                if m.category in cat_stats:
                    cat_stats[m.category]['alerte'] += 1
            
            alertes_stock.append({
                'id': m.id,
                'name': m.name,
                'category': m.category,
                'cat_label': m.get_category_display(),
                'quantity': m.quantity,
                'unit': m.unit,
                'min_threshold': m.min_threshold,
                'supplier': m.supplier.name if m.supplier else '—',
                'pct': min(pct, 100),
                'niveau': niveau,
                'icone': icone,
            })
    
    # Trier les alertes : ruptures en premier, puis critiques, puis alertes simples
    ordre_priorite = {'RUPTURE': 0, 'CRITIQUE': 1, 'ALERTE': 2}
    alertes_stock.sort(key=lambda x: (ordre_priorite.get(x['niveau'], 3), -x['pct']))
    
    # ══════════════════════════════════════════════════════════════
    # DONNÉES POUR LES GRAPHIQUES
    # ══════════════════════════════════════════════════════════════
    # Top 10 pour le graphique barres
    top_alertes = alertes_stock[:10]
    top_alertes_noms = [a['name'][:25] + '...' if len(a['name']) > 25 else a['name'] for a in top_alertes]
    top_alertes_stock = [a['quantity'] for a in top_alertes]
    top_alertes_seuil = [a['min_threshold'] for a in top_alertes]
    top_alertes_couleurs = []
    for a in top_alertes:
        if a['niveau'] == 'RUPTURE':
            top_alertes_couleurs.append('#dc2626')
        elif a['niveau'] == 'CRITIQUE':
            top_alertes_couleurs.append('#ea580c')
        else:
            top_alertes_couleurs.append('#ca8a04')
    
    # Données par catégorie pour le graphe empilé
    cat_labels = ['Film/Papier', 'Encre', 'Colle', 'Solvant']
    cat_rupture = [cat_stats['FILM']['rupture'], cat_stats['INK']['rupture'], cat_stats['GLUE']['rupture'], cat_stats['SOLV']['rupture']]
    cat_critique = [cat_stats['FILM']['critique'], cat_stats['INK']['critique'], cat_stats['GLUE']['critique'], cat_stats['SOLV']['critique']]
    cat_alerte = [cat_stats['FILM']['alerte'], cat_stats['INK']['alerte'], cat_stats['GLUE']['alerte'], cat_stats['SOLV']['alerte']]
    
    # ══════════════════════════════════════════════════════════════
    # PRÉVISIONS DE RUPTURE (seuils intelligents)
    # ══════════════════════════════════════════════════════════════
    previsions = []
    for m in all_materials:
        try:
            seuil = m.seuil_intelligent
            if seuil and seuil.consommation_journaliere_moy > 0:
                jours = seuil.jours_de_stock
                if jours <= 15:
                    previsions.append({
                        'material': m.name,
                        'stock_actuel': m.quantity,
                        'conso_jour': seuil.consommation_journaliere_moy,
                        'jours_restants': jours,
                        'date_rupture': seuil.date_rupture_prevue.strftime('%d/%m/%Y') if seuil.date_rupture_prevue else '—',
                        'critique': jours <= 7,
                    })
        except:
            pass
    previsions.sort(key=lambda x: x['jours_restants'])
    
    # ══════════════════════════════════════════════════════════════
    # AUTRES DONNÉES
    # ══════════════════════════════════════════════════════════════
    lots = StockLot.objects.select_related('material', 'fournisseur', 'emplacement').order_by('-date_reception')[:100]
    lots_bloques = StockLot.objects.filter(statut='BLOQUE').count()
    lots_attente = StockLot.objects.filter(statut='EN_ATTENTE').count()
    
    mouvements = StockMovement.objects.select_related(
        'material', 'lot', 'emplacement_source', 'emplacement_destination', 'utilisateur', 'machine', 'of'
    ).order_by('-date')[:100]
    
    locations = StockLocation.objects.filter(is_active=True).order_by('type', 'name')
    suppliers = Supplier.objects.all().order_by('name')
    
    demandes = DemandeAchat.objects.select_related('material', 'demandeur', 'valideur').order_by('-date_creation')[:50]
    da_en_attente = DemandeAchat.objects.filter(statut='SOUMISE').count()
    
    bons_commande = BonCommande.objects.select_related('fournisseur', 'cree_par').order_by('-date_commande')[:50]
    
    consos = ConsommationEncre.objects.all().order_by('-date')[:50]
    
    # Valeur totale du stock
    valeur_stock_total = sum(
        float(m.quantity) * float(m.price_per_unit) 
        for m in all_materials 
        if m.price_per_unit
    )
    
    # ══════════════════════════════════════════════════════════════
    # CONTEXTE TEMPLATE
    # ══════════════════════════════════════════════════════════════
    context = {
        # Recherche
        'search_query': search_query,
        'category_filter': category_filter,
        'low_stock_only': low_stock_only,
        'supplier_filter': supplier_filter,
        'categories': Material.CAT_CHOICES,
        
        # Matières
        'materials': materials,
        'total_matieres': Material.objects.count(),
        'suppliers': suppliers,
        
        # Alertes
        'alertes_stock': alertes_stock,
        'nb_alertes': len(alertes_stock),
        'nb_ruptures': nb_ruptures,
        'nb_critiques': nb_critiques,
        'nb_alertes_simples': nb_alertes_simples,
        
        # Graphiques (JSON)
        'top_alertes_noms': json.dumps(top_alertes_noms),
        'top_alertes_stock': json.dumps(top_alertes_stock),
        'top_alertes_seuil': json.dumps(top_alertes_seuil),
        'top_alertes_couleurs': json.dumps(top_alertes_couleurs),
        'cat_labels_json': json.dumps(cat_labels),
        'cat_rupture_json': json.dumps(cat_rupture),
        'cat_critique_json': json.dumps(cat_critique),
        'cat_alerte_json': json.dumps(cat_alerte),
        
        # Prévisions
        'previsions': previsions[:6],
        
        # Lots
        'lots': lots,
        'lots_bloques': lots_bloques,
        'lots_attente': lots_attente,
        
        # Mouvements
        'mouvements': mouvements,
        
        # Emplacements
        'locations': locations,
        
        # Achats
        'demandes': demandes,
        'da_en_attente': da_en_attente,
        'bons_commande': bons_commande,
        
        # Consos
        'consos': consos,
        
        # Stats
        'valeur_stock_total': valeur_stock_total,
    }
    
    return render(request, 'stock/stock_advanced.html', context)


@login_required
def material_search_api(request):
    """API autocomplete pour recherche dynamique AJAX"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 1:
        return JsonResponse([], safe=False)
    
    # Recherche insensible à la casse sur le nom
    materials = Material.objects.filter(
        Q(name__icontains=query) |
        Q(category__icontains=query) |
        Q(supplier__name__icontains=query)
    ).select_related('supplier').order_by('name')[:30]
    
    results = []
    for m in materials:
        results.append({
            'id': m.id,
            'name': m.name,
            'name_html': highlight_search(m.name, query),
            'category': m.get_category_display(),
            'quantity': m.quantity,
            'unit': m.unit,
            'min_threshold': m.min_threshold,
            'supplier': m.supplier.name if m.supplier else '—',
            'is_low_stock': m.is_low_stock(),
            'price': float(m.price_per_unit) if m.price_per_unit else 0,
        })
    
    return JsonResponse({
        'query': query,
        'count': len(results),
        'results': results
    })


@login_required
def export_search_results(request):
    """Export Excel des résultats de recherche"""
    query = request.GET.get('q', '').strip()
    category = request.GET.get('category', '')
    
    # Filtrer les matières
    materials = Material.objects.select_related('supplier').all()
    
    if query:
        materials = materials.filter(
            Q(name__icontains=query) |
            Q(supplier__name__icontains=query)
        )
    
    if category:
        materials = materials.filter(category=category)
    
    materials = materials.order_by('name')
    
    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Matières"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    alert_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = ['Désignation', 'Catégorie', 'Stock Actuel', 'Unité', 'Seuil Min', 'Fournisseur', 'Prix/Unité', 'Valeur Stock', 'État']
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Données
    for row_num, m in enumerate(materials, 2):
        valeur = float(m.quantity) * float(m.price_per_unit) if m.price_per_unit else 0
        etat = "⚠️ ALERTE" if m.is_low_stock() else "✓ OK"
        
        row_data = [
            m.name,
            m.get_category_display(),
            m.quantity,
            m.unit,
            m.min_threshold,
            m.supplier.name if m.supplier else '',
            float(m.price_per_unit) if m.price_per_unit else 0,
            round(valeur, 2),
            etat
        ]
        ws.append(row_data)
        
        # Appliquer le style alerte si stock faible
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if m.is_low_stock():
                cell.fill = alert_fill
    
    # Ajuster la largeur des colonnes
    column_widths = [40, 15, 15, 10, 12, 25, 12, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Préparer la réponse
    filename = f"stock_matieres_{query if query else 'all'}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ===========================================================================
# --- STOCK AVANCÉ (SUITE) ---
# ===========================================================================

@login_required
def location_list(request):
    locations = StockLocation.objects.all()
    return render(request, 'stock/stock_advanced.html', {'locations': locations})


@login_required
def location_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        type_loc = request.POST.get('type', 'GENERAL')
        desc = request.POST.get('description', '')
        if name:
            StockLocation.objects.create(name=name, type=type_loc, description=desc)
            messages.success(request, f"Emplacement « {name} » créé.")
        return redirect('stock_advanced')
    return redirect('stock_advanced')


@login_required
def location_delete(request, id):
    loc = get_object_or_404(StockLocation, id=id)
    if request.method == 'POST':
        loc.delete()
        messages.success(request, "Emplacement supprimé.")
    return redirect('stock_advanced')


@login_required
def lot_list(request):
    lots = StockLot.objects.select_related(
        'material', 'fournisseur', 'emplacement'
    ).order_by('-date_reception')
    return render(request, 'stock/stock_advanced.html', {'lots': lots})


@login_required
def lot_add(request):
    if request.method == 'POST':
        try:
            material_id = request.POST.get('material')
            numero_lot = request.POST.get('numero_lot', '').strip()
            date_reception = request.POST.get('date_reception')
            fournisseur_id = request.POST.get('fournisseur') or None
            emplacement_id = request.POST.get('emplacement') or None
            quantite = float(request.POST.get('quantite_initiale', 0))
            prix = float(request.POST.get('prix_unitaire', 0))
            statut = request.POST.get('statut', 'EN_ATTENTE')
            notes = request.POST.get('notes', '')
            material = get_object_or_404(Material, id=material_id)
            fournisseur = Supplier.objects.filter(id=fournisseur_id).first() if fournisseur_id else None
            emplacement = StockLocation.objects.filter(id=emplacement_id).first() if emplacement_id else None
            lot = StockLot.objects.create(
                material=material, numero_lot=numero_lot,
                date_reception=date_reception, fournisseur=fournisseur,
                emplacement=emplacement, quantite_initiale=quantite,
                quantite_restante=quantite, prix_unitaire=prix,
                statut=statut, notes=notes, created_by=request.user
            )
            if request.FILES.get('certificat_qualite'):
                lot.certificat_qualite = request.FILES['certificat_qualite']
                lot.save()
            if statut == 'CONFORME':
                StockMovement.objects.create(
                    type='ENTREE', material=material, lot=lot,
                    quantite=quantite, emplacement_destination=emplacement,
                    utilisateur=request.user,
                    motif=f"Réception lot {numero_lot}",
                )
            messages.success(request, f"Lot {numero_lot} créé.")
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
        return redirect('stock_advanced')
    return redirect('stock_advanced')


@login_required
def lot_valider(request, id):
    lot = get_object_or_404(StockLot, id=id)
    if request.method == 'POST':
        ancien_statut = lot.statut
        lot.statut = 'CONFORME'
        lot.save()
        if ancien_statut != 'CONFORME':
            StockMovement.objects.create(
                type='ENTREE', material=lot.material, lot=lot,
                quantite=lot.quantite_restante,
                emplacement_destination=lot.emplacement,
                utilisateur=request.user,
                motif=f"Validation lot {lot.numero_lot}",
            )
        messages.success(request, f"Lot {lot.numero_lot} validé.")
    return redirect('stock_advanced')


@login_required
def lot_bloquer(request, id):
    lot = get_object_or_404(StockLot, id=id)
    if request.method == 'POST':
        lot.statut = 'BLOQUE'
        lot.save()
        messages.warning(request, f"Lot {lot.numero_lot} bloqué.")
    return redirect('stock_advanced')


@login_required
def lot_detail(request, id):
    lot = get_object_or_404(StockLot, id=id)
    mouvements = lot.mouvements.select_related(
        'utilisateur', 'machine', 'of'
    ).order_by('-date')
    return render(request, 'stock/lot_detail.html', {
        'lot': lot,
        'mouvements': mouvements
    })


@login_required
def mouvement_add(request):
    if request.method == 'POST':
        try:
            material = get_object_or_404(Material, id=request.POST.get('material'))
            type_mvt = request.POST.get('type')
            quantite = float(request.POST.get('quantite', 0))
            lot_id = request.POST.get('lot') or None
            src_id = request.POST.get('emplacement_source') or None
            dst_id = request.POST.get('emplacement_destination') or None
            of_id = request.POST.get('of') or None
            machine_id = request.POST.get('machine') or None
            motif = request.POST.get('motif', '')
            lot = StockLot.objects.filter(id=lot_id).first() if lot_id else None
            src = StockLocation.objects.filter(id=src_id).first() if src_id else None
            dst = StockLocation.objects.filter(id=dst_id).first() if dst_id else None
            of_obj = ProductionOrder.objects.filter(id=of_id).first() if of_id else None
            machine_obj = Machine.objects.filter(id=machine_id).first() if machine_id else None
            StockMovement.objects.create(
                type=type_mvt, material=material, lot=lot,
                quantite=quantite, emplacement_source=src,
                emplacement_destination=dst, of=of_obj,
                machine=machine_obj, utilisateur=request.user, motif=motif,
            )
            messages.success(request, "Mouvement enregistré.")
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
        return redirect('stock_advanced')
    return redirect('stock_advanced')


@login_required
def da_add(request):
    if request.method == 'POST':
        try:
            import random, string
            ref = 'DA-' + ''.join(random.choices(string.digits, k=6))
            material = get_object_or_404(Material, id=request.POST.get('material'))
            DemandeAchat.objects.create(
                reference=ref, material=material,
                quantite_demandee=float(request.POST.get('quantite_demandee', 0)),
                motif=request.POST.get('motif', ''),
                urgence=request.POST.get('urgence', 'NORMALE'),
                statut='SOUMISE', demandeur=request.user,
                date_besoin=request.POST.get('date_besoin') or None,
            )
            messages.success(request, "Demande d'achat créée.")
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
        return redirect('stock_advanced')
    return redirect('stock_advanced')


@login_required
def da_valider(request, id):
    da = get_object_or_404(DemandeAchat, id=id)
    if request.method == 'POST':
        da.statut = 'VALIDEE'
        da.valideur = request.user
        da.date_validation = timezone.now().date()
        da.save()
        messages.success(request, f"DA {da.reference} validée.")
    return redirect('stock_advanced')


@login_required
def da_refuser(request, id):
    da = get_object_or_404(DemandeAchat, id=id)
    if request.method == 'POST':
        da.statut = 'REFUSEE'
        da.save()
        messages.warning(request, f"DA {da.reference} refusée.")
    return redirect('stock_advanced')


@login_required
def bc_add(request):
    if request.method == 'POST':
        try:
            import random, string
            ref = 'BC-' + ''.join(random.choices(string.digits, k=6))
            fournisseur = get_object_or_404(Supplier, id=request.POST.get('fournisseur'))
            bc = BonCommande.objects.create(
                reference=ref, fournisseur=fournisseur, statut='BROUILLON',
                date_livraison_prevue=request.POST.get('date_livraison_prevue') or None,
                notes=request.POST.get('notes', ''), cree_par=request.user,
            )
            idx = 1
            total = 0
            while request.POST.get(f'material_{idx}'):
                mat = Material.objects.filter(
                    id=request.POST.get(f'material_{idx}')
                ).first()
                if mat:
                    qte = float(request.POST.get(f'quantite_{idx}', 0))
                    prix = float(request.POST.get(f'prix_{idx}', 0))
                    LigneBonCommande.objects.create(
                        bon_commande=bc, material=mat,
                        quantite_commandee=qte, prix_unitaire=prix
                    )
                    total += qte * prix
                idx += 1
            bc.montant_total = total
            bc.save()
            messages.success(request, f"Bon de commande {bc.reference} créé.")
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
        return redirect('stock_advanced')
    return redirect('stock_advanced')


@login_required
def bc_envoyer(request, id):
    bc = get_object_or_404(BonCommande, id=id)
    if request.method == 'POST':
        bc.statut = 'ENVOYE'
        bc.save()
        messages.success(request, f"BC {bc.reference} marqué comme envoyé.")
    return redirect('stock_advanced')


@login_required
def bc_reception(request, id):
    bc = get_object_or_404(BonCommande, id=id)
    if request.method == 'POST':
        bc.statut = 'RECU_TOTAL'
        bc.date_livraison_reelle = timezone.now().date()
        bc.save()
        for ligne in bc.lignes.all():
            import random, string
            num_lot = f"LOT-{bc.reference}-{''.join(random.choices(string.digits, k=4))}"
            StockLot.objects.create(
                material=ligne.material, numero_lot=num_lot,
                date_reception=timezone.now().date(),
                fournisseur=bc.fournisseur,
                quantite_initiale=ligne.quantite_commandee,
                quantite_restante=ligne.quantite_commandee,
                prix_unitaire=ligne.prix_unitaire,
                statut='EN_ATTENTE', created_by=request.user,
            )
            ligne.quantite_recue = ligne.quantite_commandee
            ligne.save()
        messages.success(request, f"BC {bc.reference} réceptionné.")
    return redirect('stock_advanced')


@login_required
def seuil_update(request, material_id):
    material = get_object_or_404(Material, id=material_id)
    if request.method == 'POST':
        seuil, _ = StockSeuil.objects.get_or_create(
            material=material,
            defaults={
                'consommation_journaliere_moy': 0,
                'delai_fournisseur_jours': 7
            }
        )
        seuil.consommation_journaliere_moy = float(request.POST.get('conso_jour', 0))
        seuil.delai_fournisseur_jours = int(request.POST.get('delai_jours', 7))
        seuil.stock_securite_jours = int(request.POST.get('securite_jours', 3))
        seuil.save()
        messages.success(request, f"Seuil mis à jour pour {material.name}.")
    return redirect('stock_advanced')


@login_required
def stock_dashboard_data(request):
    critiques = []
    for m in Material.objects.all():
        if m.is_low_stock():
            critiques.append({
                'name': m.name,
                'stock': m.quantity,
                'seuil': m.min_threshold
            })
    return JsonResponse({'critiques': critiques})


# ===========================================================================
# --- PARC MACHINE ---
# ===========================================================================

@login_required
def machine_view(request):
    machines = Machine.objects.all().order_by('-id')
    return render(request, 'machines.html', {'machines': machines})


@login_required
def add_machine(request):
    if request.method == 'POST':
        form = MachineForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('machine_view')
    else:
        form = MachineForm()
    return render(request, 'machines.html', {'form': form, 'titre': 'Nouvelle Machine'})


# ===========================================================================
# --- MODULE PRODUCTION SPÉCIALE ---
# ===========================================================================

def _get_filtered_entries(request):
    entries = ProductionEntry.objects.all().order_by('-date', '-heure_debut')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')
    machine_id = request.GET.get('machine', '')
    support    = request.GET.get('support', '')
    equipe     = request.GET.get('equipe', '')
    produit    = request.GET.get('produit', '')

    if date_from:
        entries = entries.filter(date__gte=date_from)
    if date_to:
        entries = entries.filter(date__lte=date_to)
    if machine_id:
        entries = entries.filter(machine_id=machine_id)
    if support:
        entries = entries.filter(support=support)
    if equipe:
        entries = entries.filter(equipe=equipe)
    if produit:
        entries = entries.filter(produit=produit)
    return entries


def _get_filter_context(request):
    all_produits = ProductionEntry.objects.values_list(
        'produit', flat=True
    ).distinct().order_by('produit')

    return {
        'all_machines':       Machine.objects.all().order_by('name'),
        'all_supports':       ProductionEntry.objects.values_list(
                                  'support', flat=True
                              ).distinct().order_by('support'),
        'all_produits':       all_produits,
        'selected_date_from': request.GET.get('date_from', ''),
        'selected_date_to':   request.GET.get('date_to', ''),
        'selected_machine':   request.GET.get('machine', ''),
        'selected_support':   request.GET.get('support', ''),
        'selected_equipe':    request.GET.get('equipe', ''),
        'selected_produit':   request.GET.get('produit', ''),
    }


@login_required
def prod_dashboard(request):
    entries = _get_filtered_entries(request)
    total_prod_ml = entries.aggregate(t=Sum('prod_ml'))['t'] or 0
    total_prod_kg = entries.aggregate(t=Sum('prod_kg'))['t'] or 0
    total_dechets_kg = round(sum(e.total_dechets_kg for e in entries), 2)
    taux_dechets = round(
        (total_dechets_kg / float(total_prod_kg) * 100), 2
    ) if total_prod_kg else 0

    from collections import defaultdict
    data_par_date = defaultdict(
        lambda: {'ml': 0, 'kg': 0, 'dem': 0, 'lis': 0, 'jon': 0, 'tra': 0, 'taux': []}
    )
    for e in entries:
        d = str(e.date)
        data_par_date[d]['ml'] += e.prod_ml
        data_par_date[d]['kg'] += e.prod_kg
        data_par_date[d]['dem'] += e.dechets_demarrage
        data_par_date[d]['lis'] += e.dechets_lisiere
        data_par_date[d]['jon'] += e.dechets_jonction
        data_par_date[d]['tra'] += e.dechets_transport
        if e.prod_kg > 0:
            data_par_date[d]['taux'].append(e.taux_dechets)

    dates_sorted = sorted(data_par_date.keys())

    support_data = defaultdict(float)
    for e in entries:
        support_data[e.support] += e.prod_kg

    context = {
        'entries': entries[:20],
        'total_prod_ml': round(float(total_prod_ml), 2),
        'total_prod_kg': round(float(total_prod_kg), 2),
        'total_dechets_kg': total_dechets_kg,
        'taux_dechets': taux_dechets,
        'count': entries.count(),
        'prod_ml_dates': json.dumps(dates_sorted),
        'prod_ml_values': json.dumps([round(data_par_date[d]['ml'], 1) for d in dates_sorted]),
        'prod_kg_dates': json.dumps(dates_sorted),
        'prod_kg_values': json.dumps([round(data_par_date[d]['kg'], 2) for d in dates_sorted]),
        'prod_kg_support_labels': json.dumps(list(support_data.keys())),
        'prod_kg_support_values': json.dumps([round(v, 2) for v in support_data.values()]),
        'taux_dechets_dates': json.dumps(dates_sorted),
        'taux_dechets_values': json.dumps([
            round(sum(data_par_date[d]['taux']) / len(data_par_date[d]['taux']), 2)
            if data_par_date[d]['taux'] else 0
            for d in dates_sorted
        ]),
        'dechets_dates': json.dumps(dates_sorted),
        'dechets_dem_values': json.dumps([round(data_par_date[d]['dem'], 2) for d in dates_sorted]),
        'dechets_lis_values': json.dumps([round(data_par_date[d]['lis'], 2) for d in dates_sorted]),
        'dechets_jon_values': json.dumps([round(data_par_date[d]['jon'], 2) for d in dates_sorted]),
        'dechets_tra_values': json.dumps([round(data_par_date[d]['tra'], 2) for d in dates_sorted]),
    }
    context.update(_get_filter_context(request))
    return render(request, 'production_special/dashboard.html', context)


@login_required
def prod_saisie(request):
    if request.method == 'POST':
        form = ProductionEntryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Saisie enregistrée avec succès.")
            return redirect('prod_dashboard')
        else:
            messages.error(request, "Erreur dans le formulaire. Vérifiez les champs.")
    else:
        form = ProductionEntryForm()
    recent = ProductionEntry.objects.all().order_by('-date', '-heure_debut')[:10]
    context = {
        'form': form,
        'titre': 'Nouvelle Saisie Production',
        'recent': recent,
        'edit_mode': False,
    }
    return render(request, 'production_special/saisie.html', context)


@login_required
def prod_edit_entry(request, id):
    entry = get_object_or_404(ProductionEntry, id=id)
    if request.method == 'POST':
        form = ProductionEntryForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            messages.success(request, "Saisie mise à jour.")
            return redirect('prod_base')
        else:
            messages.error(request, "Erreur dans le formulaire.")
    else:
        form = ProductionEntryForm(instance=entry)
    recent = ProductionEntry.objects.all().order_by('-date', '-heure_debut')[:10]
    context = {
        'form': form,
        'titre': f'Modifier — {entry.produit} ({entry.date})',
        'recent': recent,
        'edit_mode': True,
        'entry': entry,
    }
    return render(request, 'production_special/saisie.html', context)


@login_required
def prod_delete_entry(request, id):
    entry = get_object_or_404(ProductionEntry, id=id)
    if request.method == 'POST':
        entry.delete()
        messages.success(request, "Saisie supprimée.")
        return redirect('prod_base')
    return render(request, 'production_special/confirm_delete.html', {'entry': entry})


@login_required
def prod_base(request):
    entries = _get_filtered_entries(request)
    context = {'entries': entries}
    context.update(_get_filter_context(request))
    return render(request, 'production_special/base.html', context)


@login_required
def prod_detail_qualite(request):
    entries = _get_filtered_entries(request)
    total_dechets_kg = round(sum(e.total_dechets_kg for e in entries), 2)
    total_dechets_demarrage = round(sum(e.dechets_demarrage for e in entries), 2)
    total_dechets_lisiere = round(sum(e.dechets_lisiere for e in entries), 2)
    total_dechets_jonction = round(sum(e.dechets_jonction for e in entries), 2)
    total_dechets_transport = round(sum(e.dechets_transport for e in entries), 2)

    from collections import defaultdict
    data_dates = defaultdict(lambda: {
        'dem_a': 0, 'dem_b': 0, 'lis_a': 0, 'lis_b': 0,
        'jon_a': 0, 'jon_b': 0, 'tra_a': 0, 'tra_b': 0, 'total': 0
    })
    for e in entries:
        d = str(e.date)
        eq = e.equipe
        data_dates[d]['dem_' + eq.lower()] += e.dechets_demarrage
        data_dates[d]['lis_' + eq.lower()] += e.dechets_lisiere
        data_dates[d]['jon_' + eq.lower()] += e.dechets_jonction
        data_dates[d]['tra_' + eq.lower()] += e.dechets_transport
        data_dates[d]['total'] += e.total_dechets_kg

    dates_sorted = sorted(data_dates.keys())
    context = {
        'entries': entries,
        'total_dechets_kg': total_dechets_kg,
        'total_dechets_demarrage': total_dechets_demarrage,
        'total_dechets_lisiere': total_dechets_lisiere,
        'total_dechets_jonction': total_dechets_jonction,
        'total_dechets_transport': total_dechets_transport,
        'dates_labels': json.dumps(dates_sorted),
        'dem_a': json.dumps([round(data_dates[d]['dem_a'], 2) for d in dates_sorted]),
        'dem_b': json.dumps([round(data_dates[d]['dem_b'], 2) for d in dates_sorted]),
        'lis_a': json.dumps([round(data_dates[d]['lis_a'], 2) for d in dates_sorted]),
        'lis_b': json.dumps([round(data_dates[d]['lis_b'], 2) for d in dates_sorted]),
        'jon_a': json.dumps([round(data_dates[d]['jon_a'], 2) for d in dates_sorted]),
        'jon_b': json.dumps([round(data_dates[d]['jon_b'], 2) for d in dates_sorted]),
        'tra_a': json.dumps([round(data_dates[d]['tra_a'], 2) for d in dates_sorted]),
        'tra_b': json.dumps([round(data_dates[d]['tra_b'], 2) for d in dates_sorted]),
        'h_dates': json.dumps(dates_sorted),
        'h_dem': json.dumps([round(data_dates[d]['dem_a'] + data_dates[d]['dem_b'], 2) for d in dates_sorted]),
        'h_lis': json.dumps([round(data_dates[d]['lis_a'] + data_dates[d]['lis_b'], 2) for d in dates_sorted]),
        'h_jon': json.dumps([round(data_dates[d]['jon_a'] + data_dates[d]['jon_b'], 2) for d in dates_sorted]),
        'h_tra': json.dumps([round(data_dates[d]['tra_a'] + data_dates[d]['tra_b'], 2) for d in dates_sorted]),
    }
    context.update(_get_filter_context(request))
    return render(request, 'production_special/detail_qualite.html', context)


@login_required
def prod_synthese_temps(request):
    entries = _get_filtered_entries(request)
    decalage_total = round(sum(e.decalage for e in entries), 2)
    temps_ouverture_total_min = sum(e.temps_ouverture_minutes for e in entries)
    heures = int(temps_ouverture_total_min // 60)
    minutes = int(temps_ouverture_total_min % 60)
    temps_ouverture_total = f"{heures}:{minutes:02d}"
    total_rebobinage = round(sum(e.rebobinage_kg for e in entries), 2)

    from collections import defaultdict
    data_dec = defaultdict(float)
    data_temps_a = defaultdict(float)
    data_temps_b = defaultdict(float)
    produits_set = set()

    for e in entries:
        d = str(e.date)
        data_dec[d] += e.decalage
        produits_set.add(e.produit[:20])
        if e.equipe == 'A':
            data_temps_a[e.produit[:20]] += e.temps_ouverture_minutes / 60
        elif e.equipe == 'B':
            data_temps_b[e.produit[:20]] += e.temps_ouverture_minutes / 60

    dates_sorted = sorted(data_dec.keys())
    produits_sorted = sorted(produits_set)

    context = {
        'entries': entries,
        'decalage_total': decalage_total,
        'temps_ouverture_total': temps_ouverture_total,
        'total_rebobinage': total_rebobinage,
        'decalage_dates': json.dumps(dates_sorted),
        'decalage_values': json.dumps([round(data_dec[d], 2) for d in dates_sorted]),
        'temps_produit_labels': json.dumps(produits_sorted),
        'temps_produit_a': json.dumps([round(data_temps_a.get(p, 0), 2) for p in produits_sorted]),
        'temps_produit_b': json.dumps([round(data_temps_b.get(p, 0), 2) for p in produits_sorted]),
    }
    context.update(_get_filter_context(request))
    return render(request, 'production_special/synthese_temps.html', context)


# ===========================================================================
# --- MODULE PRODUCTION SPÉCIAL — CONSOMMATION ENCRE & DÉCHETS IMPRESSION ---
# ===========================================================================

def _get_filtered_encre(request):
    """Filtre les enregistrements ConsommationEncre selon les paramètres GET"""
    consos = ConsommationEncre.objects.all().order_by('-date')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    process   = request.GET.get('process_type', '')
    support   = request.GET.get('support', '')
    job       = request.GET.get('job', '')

    if date_from:
        consos = consos.filter(date__gte=date_from)
    if date_to:
        consos = consos.filter(date__lte=date_to)
    if process:
        consos = consos.filter(process_type=process)
    if support:
        consos = consos.filter(support__icontains=support)
    if job:
        consos = consos.filter(job_name=job)
    return consos


def _get_encre_filter_context(request):
    """Contexte commun pour les filtres encre"""
    all_supports = ConsommationEncre.objects.values_list(
        'support', flat=True
    ).distinct().order_by('support')
    
    all_jobs = ConsommationEncre.objects.values_list(
        'job_name', flat=True
    ).distinct().order_by('job_name')
    
    return {
        'all_supports_encre': all_supports,
        'all_jobs_encre':     list(all_jobs),
        'process_choices':    ConsommationEncre.PROCESS_CHOICES,
        'sel_date_from':      request.GET.get('date_from', ''),
        'sel_date_to':        request.GET.get('date_to', ''),
        'sel_process':        request.GET.get('process_type', ''),
        'sel_support':        request.GET.get('support', ''),
        'sel_job':            request.GET.get('job', ''),
    }


@login_required
def encre_dashboard(request):
    """Tableau de bord principal — Consommation Encre & Déchets Impression"""
    from collections import defaultdict

    consos = _get_filtered_encre(request)
    count  = consos.count()

    total_encre   = sum(c.total_encre   for c in consos)
    total_solvant = sum(c.total_solvant for c in consos)
    total_injecte = total_encre + total_solvant

    total_gain_masse    = sum(c.gain_de_masse_kg    for c in consos)
    total_evaporee      = sum(c.matiere_evaporee_kg for c in consos)

    taux_gain_global = round(
        (total_gain_masse / total_injecte * 100), 2
    ) if total_injecte else 0
    taux_evap_global = round(
        (total_evaporee / total_injecte * 100), 2
    ) if total_injecte else 0

    total_metrage = sum(c.metrage for c in consos)

    grammages = [c.grammage for c in consos if c.grammage > 0]
    grammage_moyen = round(sum(grammages) / len(grammages), 2) if grammages else 0

    couleurs_noms = [
        'Noir', 'Magenta', 'Jaune', 'Cyan',
        'Doré', 'Silver', 'Orange', 'Blanc', 'Vernis'
    ]
    couleurs_vals = [
        round(consos.aggregate(t=Sum('encre_noir'))['t']    or 0, 2),
        round(consos.aggregate(t=Sum('encre_magenta'))['t'] or 0, 2),
        round(consos.aggregate(t=Sum('encre_jaune'))['t']   or 0, 2),
        round(consos.aggregate(t=Sum('encre_cyan'))['t']    or 0, 2),
        round(consos.aggregate(t=Sum('encre_dore'))['t']    or 0, 2),
        round(consos.aggregate(t=Sum('encre_silver'))['t']  or 0, 2),
        round(consos.aggregate(t=Sum('encre_orange'))['t']  or 0, 2),
        round(consos.aggregate(t=Sum('encre_blanc'))['t']   or 0, 2),
        round(consos.aggregate(t=Sum('encre_vernis'))['t']  or 0, 2),
    ]
    couleurs_colors = [
        '#1a1a1a', '#e91e8c', '#ffd600', '#00b8d9',
        '#ffc107', '#9e9e9e', '#ff6d00', '#f5f5f5', '#4caf50'
    ]

    data_dates = defaultdict(lambda: {
        'encre': 0, 'solvant': 0, 'gain': 0,
        'evap': 0, 'metrage': 0
    })
    for c in consos:
        d = str(c.date)
        data_dates[d]['encre']   += c.total_encre
        data_dates[d]['solvant'] += c.total_solvant
        data_dates[d]['gain']    += c.gain_de_masse_kg
        data_dates[d]['evap']    += c.matiere_evaporee_kg
        data_dates[d]['metrage'] += c.metrage

    dates_sorted = sorted(data_dates.keys())

    pie_labels = ['Encres', 'Solvants']
    pie_vals   = [round(total_encre, 2), round(total_solvant, 2)]
    pie_colors = ['#f3b83a', '#38bdf8']

    pie2_labels = ['Gain de masse', 'Matière évaporée']
    pie2_vals   = [round(total_gain_masse, 2), round(total_evaporee, 2)]
    pie2_colors = ['#22c55e', '#ef4444']

    job_data = defaultdict(float)
    for c in consos:
        job_data[c.job_name[:30]] += c.total_encre + c.total_solvant
    top_jobs = sorted(job_data.items(), key=lambda x: x[1], reverse=True)[:8]

    flexo_count = consos.filter(process_type='FLEXO').count()
    helio_count = consos.filter(process_type='HELIO').count()

    couleurs_table = [
        {'nom': n, 'val': v, 'color': c}
        for n, v, c in zip(couleurs_noms, couleurs_vals, couleurs_colors)
        if v > 0
    ]

    context = {
        'count':            count,
        'total_encre':      round(total_encre, 2),
        'total_solvant':    round(total_solvant, 2),
        'total_injecte':    round(total_injecte, 2),
        'total_gain_masse': round(total_gain_masse, 2),
        'total_evaporee':   round(total_evaporee, 2),
        'taux_gain_global': taux_gain_global,
        'taux_evap_global': taux_evap_global,
        'total_metrage':    round(total_metrage, 2),
        'grammage_moyen':   grammage_moyen,
        'flexo_count':      flexo_count,
        'helio_count':      helio_count,
        'couleurs_table':   couleurs_table,
        'consos_recentes':  consos[:15],

        'chart_couleurs_labels': json.dumps(couleurs_noms),
        'chart_couleurs_vals':   json.dumps(couleurs_vals),
        'chart_couleurs_colors': json.dumps(couleurs_colors),

        'chart_dates':       json.dumps(dates_sorted),
        'chart_encre_vals':  json.dumps(
            [round(data_dates[d]['encre'], 2) for d in dates_sorted]
        ),
        'chart_solvant_vals': json.dumps(
            [round(data_dates[d]['solvant'], 2) for d in dates_sorted]
        ),
        'chart_gain_vals':    json.dumps(
            [round(data_dates[d]['gain'], 2) for d in dates_sorted]
        ),
        'chart_evap_vals':    json.dumps(
            [round(data_dates[d]['evap'], 2) for d in dates_sorted]
        ),
        'chart_metrage_vals': json.dumps(
            [round(data_dates[d]['metrage'], 1) for d in dates_sorted]
        ),

        'pie_labels':  json.dumps(pie_labels),
        'pie_vals':    json.dumps(pie_vals),
        'pie_colors':  json.dumps(pie_colors),

        'pie2_labels': json.dumps(pie2_labels),
        'pie2_vals':   json.dumps(pie2_vals),
        'pie2_colors': json.dumps(pie2_colors),

        'chart_jobs_labels': json.dumps([j[0] for j in top_jobs]),
        'chart_jobs_vals':   json.dumps([round(j[1], 2) for j in top_jobs]),

        'chart_process_labels': json.dumps(['Flexo', 'Hélio']),
        'chart_process_vals':   json.dumps([flexo_count, helio_count]),
    }
    context.update(_get_encre_filter_context(request))
    return render(request, 'production_special/encre_dashboard.html', context)


@login_required
def encre_saisie(request):
    """Formulaire de saisie consommation encre"""
    if request.method == 'POST':
        form = ConsommationEncreForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Consommation encre enregistrée avec succès.")
            return redirect('encre_dashboard')
        else:
            messages.error(request, "❌ Erreur dans le formulaire. Vérifiez les champs.")
    else:
        form = ConsommationEncreForm()

    recent = ConsommationEncre.objects.all().order_by('-date')[:10]
    context = {
        'form':      form,
        'titre':     'Nouvelle Saisie — Encre & Solvants',
        'recent':    recent,
        'edit_mode': False,
    }
    return render(request, 'production_special/encre_saisie.html', context)


@login_required
def encre_edit(request, id):
    """Modifier une saisie consommation encre"""
    conso = get_object_or_404(ConsommationEncre, id=id)
    if request.method == 'POST':
        form = ConsommationEncreForm(request.POST, instance=conso)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Saisie mise à jour.")
            return redirect('encre_dashboard')
        else:
            messages.error(request, "❌ Erreur dans le formulaire.")
    else:
        form = ConsommationEncreForm(instance=conso)

    recent = ConsommationEncre.objects.all().order_by('-date')[:10]
    context = {
        'form':      form,
        'titre':     f'Modifier — {conso.job_name} ({conso.date})',
        'recent':    recent,
        'edit_mode': True,
        'conso':     conso,
    }
    return render(request, 'production_special/encre_saisie.html', context)


@login_required
def encre_delete(request, id):
    """Supprimer une saisie consommation encre"""
    conso = get_object_or_404(ConsommationEncre, id=id)
    if request.method == 'POST':
        conso.delete()
        messages.success(request, "🗑️ Saisie supprimée.")
        return redirect('encre_dashboard')
    return render(
        request,
        'production_special/encre_confirm_delete.html',
        {'conso': conso}
    )


@login_required
def encre_detail(request, id):
    """Détail d'une saisie avec calculs complets"""
    conso = get_object_or_404(ConsommationEncre, id=id)

    couleurs_detail = [
        {'nom': 'Noir',    'val': conso.encre_noir,    'color': '#1a1a1a', 'bg': 'bg-gray-800'},
        {'nom': 'Magenta', 'val': conso.encre_magenta, 'color': '#e91e8c', 'bg': 'bg-pink-900'},
        {'nom': 'Jaune',   'val': conso.encre_jaune,   'color': '#ffd600', 'bg': 'bg-yellow-800'},
        {'nom': 'Cyan',    'val': conso.encre_cyan,    'color': '#00b8d9', 'bg': 'bg-cyan-900'},
        {'nom': 'Doré',    'val': conso.encre_dore,    'color': '#ffc107', 'bg': 'bg-amber-800'},
        {'nom': 'Silver',  'val': conso.encre_silver,  'color': '#9e9e9e', 'bg': 'bg-slate-600'},
        {'nom': 'Orange',  'val': conso.encre_orange,  'color': '#ff6d00', 'bg': 'bg-orange-900'},
        {'nom': 'Blanc',   'val': conso.encre_blanc,   'color': '#f5f5f5', 'bg': 'bg-slate-500'},
        {'nom': 'Vernis',  'val': conso.encre_vernis,  'color': '#4caf50', 'bg': 'bg-green-900'},
    ]
    couleurs_detail = [c for c in couleurs_detail if c['val'] > 0]

    pie_noms   = [c['nom']   for c in couleurs_detail]
    pie_vals_d = [c['val']   for c in couleurs_detail]
    pie_cols_d = [c['color'] for c in couleurs_detail]

    context = {
        'conso':          conso,
        'couleurs_detail': couleurs_detail,
        'pie_noms_json':  json.dumps(pie_noms),
        'pie_vals_json':  json.dumps(pie_vals_d),
        'pie_cols_json':  json.dumps(pie_cols_d),
    }
    return render(request, 'production_special/encre_detail.html', context)


@login_required
def encre_analyse(request):
    """Analyse avancée — Déchets impression, grammage, comparaison Flexo/Hélio"""
    from collections import defaultdict

    consos  = _get_filtered_encre(request)
    count   = consos.count()

    dechet_dates   = defaultdict(float)
    dechet_flexo   = defaultdict(float)
    dechet_helio   = defaultdict(float)
    grammage_dates = defaultdict(list)

    for c in consos:
        d = str(c.date)
        dechet_dates[d]   += c.matiere_evaporee_kg
        grammage_dates[d].append(c.grammage)
        if c.process_type == 'FLEXO':
            dechet_flexo[d] += c.matiere_evaporee_kg
        else:
            dechet_helio[d] += c.matiere_evaporee_kg

    dates_sorted = sorted(dechet_dates.keys())

    flexo_qs = consos.filter(process_type='FLEXO')
    helio_qs = consos.filter(process_type='HELIO')

    def _stats(qs):
        total_inj = sum(c.total_encre + c.total_solvant for c in qs)
        total_evp = sum(c.matiere_evaporee_kg for c in qs)
        total_gai = sum(c.gain_de_masse_kg for c in qs)
        gram_list = [c.grammage for c in qs if c.grammage > 0]
        return {
            'count':       qs.count(),
            'total_inj':   round(total_inj, 2),
            'total_evp':   round(total_evp, 2),
            'total_gai':   round(total_gai, 2),
            'taux_evp':    round(total_evp / total_inj * 100, 2) if total_inj else 0,
            'gram_moy':    round(sum(gram_list) / len(gram_list), 2) if gram_list else 0,
        }

    stats_flexo = _stats(flexo_qs)
    stats_helio = _stats(helio_qs)

    radar_labels = [
        'Total Injecté', 'Total Évaporé',
        'Gain Masse', 'Taux Évap %', 'Grammage moy.'
    ]
    radar_flexo = [
        stats_flexo['total_inj'],  stats_flexo['total_evp'],
        stats_flexo['total_gai'],  stats_flexo['taux_evp'],
        stats_flexo['gram_moy'],
    ]
    radar_helio = [
        stats_helio['total_inj'],  stats_helio['total_evp'],
        stats_helio['total_gai'],  stats_helio['taux_evp'],
        stats_helio['gram_moy'],
    ]

    support_data = defaultdict(float)
    for c in consos:
        support_data[c.support[:25]] += c.total_encre

    top_support = sorted(support_data.items(), key=lambda x: x[1], reverse=True)[:10]

    job_evap = []
    job_seen = {}
    for c in consos:
        key = c.job_name[:30]
        if key not in job_seen:
            job_seen[key] = {'evap': 0, 'inj': 0}
        job_seen[key]['evap'] += c.matiere_evaporee_kg
        job_seen[key]['inj']  += c.total_encre + c.total_solvant

    for job, vals in sorted(
        job_seen.items(),
        key=lambda x: x[1]['evap'],
        reverse=True
    )[:8]:
        taux = round(vals['evap'] / vals['inj'] * 100, 1) if vals['inj'] else 0
        job_evap.append({'job': job, 'evap': round(vals['evap'], 2), 'taux': taux})

    context = {
        'count':        count,
        'stats_flexo':  stats_flexo,
        'stats_helio':  stats_helio,
        'job_evap':     job_evap,

        'chart_dec_dates':  json.dumps(dates_sorted),
        'chart_dec_total':  json.dumps(
            [round(dechet_dates[d], 2) for d in dates_sorted]
        ),
        'chart_dec_flexo':  json.dumps(
            [round(dechet_flexo.get(d, 0), 2) for d in dates_sorted]
        ),
        'chart_dec_helio':  json.dumps(
            [round(dechet_helio.get(d, 0), 2) for d in dates_sorted]
        ),

        'chart_gram_dates': json.dumps(dates_sorted),
        'chart_gram_vals':  json.dumps([
            round(
                sum(grammage_dates[d]) / len(grammage_dates[d]), 2
            ) if grammage_dates[d] else 0
            for d in dates_sorted
        ]),

        'radar_labels':     json.dumps(radar_labels),
        'radar_flexo':      json.dumps(radar_flexo),
        'radar_helio':      json.dumps(radar_helio),

        'chart_sup_labels': json.dumps([s[0] for s in top_support]),
        'chart_sup_vals':   json.dumps([round(s[1], 2) for s in top_support]),

        'chart_job_labels': json.dumps([j['job']  for j in job_evap]),
        'chart_job_evap':   json.dumps([j['evap'] for j in job_evap]),
        'chart_job_taux':   json.dumps([j['taux'] for j in job_evap]),
    }
    context.update(_get_encre_filter_context(request))
    return render(request, 'production_special/encre_analyse.html', context)


# ===========================================================================
# --- IMPORT EXCEL ---
# ===========================================================================

def parse_time_safe(val):
    if val is None or val == 0 or val == '' or str(val).strip() == '':
        return None
    try:
        if hasattr(val, 'hour'):
            return val
        if hasattr(val, 'time'):
            return val.time()
        if isinstance(val, float):
            total_seconds = int(val * 24 * 3600)
            return datetime.time(total_seconds // 3600, (total_seconds % 3600) // 60)
        s = str(val).strip()
        parts = s.replace('h', ':').replace('H', ':').split(':')
        if len(parts) >= 2:
            return datetime.time(int(parts[0]), int(parts[1]))
        return None
    except Exception:
        return None


@login_required
def import_stock_view(request):
    context = {}
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            import_type = request.POST.get('import_type')
            excel_file = request.FILES['excel_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)
            df = pd.read_excel(file_path).fillna(0)
            count = 0
            errors = 0
            details = []

            if import_type == 'STOCK':
                for idx, row in df.iterrows():
                    try:
                        cat_map = {'Film': 'FILM', 'Encre': 'INK', 'Colle': 'GLUE', 'Solvant': 'SOLV'}
                        cat_val = row.get('Categorie', 'FILM')
                        if cat_val == 0:
                            cat_val = 'FILM'
                        Material.objects.update_or_create(
                            name=row.get('Designation', 'Inconnu'),
                            defaults={
                                'category': cat_map.get(cat_val, 'FILM'),
                                'quantity': row.get('Quantite', 0),
                                'unit': row.get('Unite', 'kg'),
                                'min_threshold': row.get('Seuil_Min', 0),
                                'price_per_unit': row.get('Prix', 0)
                            }
                        )
                        count += 1
                        details.append(f"Ligne {idx+2}: ✅ {row.get('Designation')} importé")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ {str(e)}")

            elif import_type == 'CRM':
                for idx, row in df.iterrows():
                    try:
                        nom = row.get('Nom')
                        if nom and nom != 0:
                            status_map = {
                                'Active': 'ACTIVE',
                                'Prospect': 'PROSPECT',
                                'VIP': 'VIP'
                            }
                            Client.objects.update_or_create(
                                name=nom,
                                defaults={
                                    'city': row.get('Ville', ''),
                                    'phone': row.get('Telephone', ''),
                                    'email': row.get('Email', ''),
                                    'sector': row.get('Secteur', ''),
                                    'status': status_map.get(row.get('Statut'), 'PROSPECT')
                                }
                            )
                            count += 1
                            details.append(f"Ligne {idx+2}: ✅ Client {nom} importé")
                        else:
                            details.append(f"Ligne {idx+2}: ⚠️ Nom vide — ignorée")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ {str(e)}")

            elif import_type == 'SPECIAL_PROD':
                for idx, row in df.iterrows():
                    try:
                        try:
                            d_val = pd.to_datetime(row.get('Date')).date()
                        except Exception:
                            d_val = datetime.date.today()

                        produit_val = str(row.get('Produit', '')).strip()
                        if not produit_val or produit_val == '0':
                            details.append(f"Ligne {idx+2}: ⚠️ Produit vide — ignorée")
                            continue

                        client_obj = None
                        cli_name = str(row.get('Client', '')).strip()
                        if cli_name and cli_name not in ('0', ''):
                            client_obj, _ = Client.objects.get_or_create(
                                name=cli_name,
                                defaults={'city': 'Non renseigné', 'phone': '000'}
                            )

                        machine_obj = None
                        mac_name = str(row.get('Machine', '')).strip()
                        if mac_name and mac_name not in ('0', ''):
                            machine_obj = Machine.objects.filter(
                                name__icontains=mac_name
                            ).first()
                            if not machine_obj:
                                machine_obj = Machine.objects.create(
                                    name=mac_name, type='IMP', status='STOP'
                                )
                                details.append(f"  → Machine '{mac_name}' créée automatiquement")

                        equipe_val = str(row.get('Equipe', 'A')).strip().upper()
                        if equipe_val not in ['A', 'B', 'C']:
                            equipe_val = 'A'

                        h_debut = parse_time_safe(row.get('H_Debut'))
                        h_fin = parse_time_safe(row.get('H_Fin'))

                        def safe_float(v, d=0):
                            try:
                                x = float(v)
                                return x if x == x else d
                            except Exception:
                                return d

                        entry = ProductionEntry.objects.create(
                            date=d_val,
                            produit=produit_val,
                            support=str(row.get('Support', '')),
                            quantite_lancee=safe_float(row.get('Qte_Lancee')),
                            lot=str(row.get('Lot', '')),
                            laize=safe_float(row.get('Laize')),
                            client=client_obj,
                            equipe=equipe_val,
                            machine=machine_obj,
                            heure_debut=h_debut,
                            heure_fin=h_fin,
                            prod_ml=safe_float(row.get('Prod_ML')),
                            dechets_demarrage=safe_float(row.get('Dec_Demarrage')),
                            dechets_lisiere=safe_float(row.get('Dec_Lisiere')),
                            dechets_jonction=safe_float(row.get('Dec_Jonction')),
                            dechets_transport=safe_float(row.get('Dec_Transport')),
                            prod_kg=safe_float(row.get('Prod_KG')),
                            rebobinage_kg=safe_float(row.get('Rebobinage_KG'))
                        )
                        count += 1
                        details.append(
                            f"Ligne {idx+2}: ✅ {produit_val} du {d_val} importé (ID={entry.id})"
                        )
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ ERREUR — {str(e)}")

            elif import_type == 'CONSO':
                for idx, row in df.iterrows():
                    try:
                        try:
                            d_prod = pd.to_datetime(row.get('Date')).date()
                        except Exception:
                            d_prod = datetime.date.today()
                        raw_type = str(row.get('Type', 'FLEXO')).upper()
                        final_type = 'HELIO' if 'HELIO' in raw_type else 'FLEXO'
                        ConsommationEncre.objects.create(
                            date=d_prod, process_type=final_type,
                            support=row.get('Support', 'Inconnu'),
                            laize=float(row.get('Laize', 0)),
                            bobine_in=float(row.get('Bobine_In', 0)),
                            bobine_out=float(row.get('Bobine_Out', 0)),
                            metrage=float(row.get('Metrage', 0)),
                            encre_noir=float(row.get('Noir', 0)),
                            encre_magenta=float(row.get('Magenta', 0)),
                            encre_jaune=float(row.get('Jaune', 0)),
                            encre_cyan=float(row.get('Cyan', 0)),
                            encre_dore=float(row.get('Dore', 0)),
                            encre_silver=float(row.get('Silver', 0)),
                            encre_orange=float(row.get('Orange', 0)),
                            encre_blanc=float(row.get('Blanc', 0)),
                            encre_vernis=float(row.get('Vernis', 0)),
                            solvant_metoxyn=float(row.get('Metoxyn', 0)),
                            solvant_2080=float(row.get('2080', 0))
                        )
                        count += 1
                        details.append(f"Ligne {idx+2}: ✅ Conso {d_prod} importée")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ {str(e)}")

            elif import_type == 'TOOLS':
                for idx, row in df.iterrows():
                    try:
                        prod_ref = row.get('Ref_Produit')
                        if prod_ref and prod_ref != 0:
                            cli, _ = Client.objects.get_or_create(
                                name="Client Divers",
                                defaults={'city': 'Interne', 'phone': '000'}
                            )
                            product, _ = TechnicalProduct.objects.get_or_create(
                                ref_internal=prod_ref,
                                defaults={
                                    'name': f"Produit {prod_ref}",
                                    'client': cli,
                                    'structure_type': 'MONO',
                                    'width_mm': 100
                                }
                            )
                            type_map = {'Cylindre': 'CYL', 'Cliche': 'CLICHE'}
                            type_val = row.get('Type', 'CYL')
                            if type_val == 0:
                                type_val = 'CYL'
                            Tooling.objects.update_or_create(
                                serial_number=row.get('Serial'),
                                defaults={
                                    'product': product,
                                    'tool_type': type_map.get(type_val, 'CYL'),
                                    'max_impressions': row.get('Tours_Max', 1000000),
                                    'current_impressions': row.get('Tours_Actuels', 0)
                                }
                            )
                            count += 1
                            details.append(f"Ligne {idx+2}: ✅ Outil {row.get('Serial')} importé")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ {str(e)}")

            elif import_type == 'PLANNING':
                for idx, row in df.iterrows():
                    try:
                        cli_name = row.get('Client')
                        if cli_name and cli_name != 0:
                            client, _ = Client.objects.get_or_create(
                                name=cli_name,
                                defaults={'city': '?', 'phone': '?'}
                            )
                            prod_name = row.get('Produit')
                            product, _ = TechnicalProduct.objects.get_or_create(
                                ref_internal=f"REF-{str(prod_name)[:5]}",
                                defaults={
                                    'name': prod_name,
                                    'client': client,
                                    'structure_type': 'MONO',
                                    'width_mm': 500
                                }
                            )
                            mac_name = row.get('Machine')
                            machine = Machine.objects.filter(
                                name__icontains=str(mac_name)
                            ).first()
                            try:
                                start_d = pd.to_datetime(row.get('Date_Debut'))
                            except Exception:
                                start_d = datetime.datetime.now()
                            end_d = start_d + timedelta(hours=4)
                            ProductionOrder.objects.update_or_create(
                                of_number=str(row.get('OF_Numero')),
                                defaults={
                                    'client': client,
                                    'product': product,
                                    'machine': machine,
                                    'start_time': start_d,
                                    'end_time': end_d,
                                    'quantity_planned': row.get('Qte_Prevue', 0),
                                    'status': 'PLANNED'
                                }
                            )
                            count += 1
                            details.append(f"Ligne {idx+2}: ✅ OF {row.get('OF_Numero')} importé")
                    except Exception as e:
                        errors += 1
                        details.append(f"Ligne {idx+2}: ❌ {str(e)}")

            context = {
                'message': (
                    f'⚠️ {count} OK, {errors} erreurs.'
                    if errors else
                    f'✅ {count} lignes importées avec succès !'
                ),
                'success': count > 0,
                'details': details
            }
            try:
                os.remove(file_path)
            except Exception:
                pass

        except Exception as e:
            context = {'message': f'❌ Erreur critique : {str(e)}', 'success': False}

    return render(request, 'stock/import_stock.html', context)


def download_template_special_prod(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Speciale"
    headers = [
        'Date', 'Produit', 'Support', 'Qte_Lancee', 'Lot', 'Laize',
        'Client', 'Equipe', 'Machine', 'H_Debut', 'H_Fin', 'Prod_ML',
        'Dec_Demarrage', 'Dec_Lisiere', 'Dec_Jonction', 'Dec_Transport',
        'Prod_KG', 'Rebobinage_KG'
    ]
    hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
    tb = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        cell.border = tb
    example = [
        '15/01/2025', 'Sac Lait 1L', 'PEBD 50μ', 500, 'LOT-001', 320,
        'Laiterie Atlas', 'A', 'IMP-01', '08:00', '16:30', 12000,
        5.2, 3.1, 1.5, 2.0, 485, 10
    ]
    ef = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = ef
        cell.border = tb
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = ml + 4
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="Template_Production_Speciale.xlsx"'
    )
    wb.save(response)
    return response


# ===========================================================================
# --- ADMINISTRATION PERSONNALISÉE ---
# ===========================================================================

@login_required
@staff_member_required
def admin_view(request):
    users = User.objects.all().order_by('-date_joined')
    total_users = users.count()
    active_users = users.filter(is_active=True).count()
    staff_users = users.filter(is_staff=True).count()
    groups = Group.objects.annotate(user_count=Count('user')).all()

    stats_modules = [
        {
            'nom': 'CRM & Clients',
            'icone': '🤝',
            'items': [
                {'label': 'Clients', 'count': Client.objects.count(), 'url': 'crm_view'},
                {'label': 'Opportunités', 'count': Opportunite.objects.count(), 'url': 'opportunites_view'},
                {'label': 'Devis', 'count': Quote.objects.count(), 'url': 'quotes_view'},
            ]
        },
        {
            'nom': 'Production',
            'icone': '🏭',
            'items': [
                {'label': 'OF Multi-Processus', 'count': OrdreFabrication.objects.count(), 'url': 'of_list'},
                {'label': 'Saisies Production', 'count': ProductionEntry.objects.count(), 'url': 'prod_dashboard'},
            ]
        },
        {
            'nom': 'Stock & Machines',
            'icone': '📦',
            'items': [
                {'label': 'Matières Premières', 'count': Material.objects.count(), 'url': 'stock_advanced'},
                {'label': 'Machines', 'count': Machine.objects.count(), 'url': 'machine_view'},
            ]
        },
    ]

    recent_actions = LogEntry.objects.select_related(
        'user', 'content_type'
    ).order_by('-action_time')[:20]

    context = {
        'users': users,
        'total_users': total_users,
        'active_users': active_users,
        'staff_users': staff_users,
        'groups': groups,
        'stats_modules': stats_modules,
        'recent_actions': recent_actions,
        'page_title': 'Administration',
    }
    return render(request, 'admin_custom.html', context)


@login_required
@staff_member_required
def admin_add_user(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        is_staff = request.POST.get('is_staff') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        if username and password:
            if not User.objects.filter(username=username).exists():
                u = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password
                )
                u.is_staff = is_staff
                u.is_active = is_active
                u.save()
                messages.success(request, f"✅ Utilisateur '{username}' créé avec succès !")
            else:
                messages.error(request, f"❌ Le nom d'utilisateur '{username}' existe déjà.")
        else:
            messages.error(request, "❌ Nom d'utilisateur et mot de passe obligatoires.")
    return redirect('admin_view')


@login_required
@staff_member_required
def admin_edit_user(request, user_id):
    u = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        is_staff = request.POST.get('is_staff') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        u.username = username or u.username
        u.email = email
        u.is_staff = is_staff
        u.is_active = is_active
        if password:
            u.set_password(password)
        u.save()
        messages.success(request, f"✅ Utilisateur '{u.username}' mis à jour !")
    return redirect('admin_view')


@login_required
@staff_member_required
def admin_toggle_user(request, user_id):
    u = get_object_or_404(User, id=user_id)
    if not u.is_superuser:
        u.is_active = not u.is_active
        u.save()
        etat = "activé" if u.is_active else "désactivé"
        messages.success(request, f"✅ Utilisateur '{u.username}' {etat}.")
    return redirect('admin_view')


# ===========================================================================
# --- CHAT EN TEMPS RÉEL ---
# ===========================================================================

@login_required
def chat_home(request):
    """Page principale du chat avec liste des salons"""
    rooms = ChatRoom.objects.filter(est_actif=True).order_by('type', 'name')
    
    default_rooms = [
        {'name': 'Général', 'slug': 'general', 'type': 'GENERAL', 'icone': '💬'},
        {'name': 'Production', 'slug': 'production', 'type': 'PRODUCTION', 'icone': '🏭'},
        {'name': 'Commercial', 'slug': 'commercial', 'type': 'COMMERCIAL', 'icone': '💼'},
        {'name': 'Technique', 'slug': 'technique', 'type': 'TECHNIQUE', 'icone': '🔧'},
        {'name': 'Urgences', 'slug': 'urgences', 'type': 'URGENCE', 'icone': '🚨'},
    ]
    
    for room_data in default_rooms:
        ChatRoom.objects.get_or_create(
            slug=room_data['slug'],
            defaults=room_data
        )
    
    rooms = ChatRoom.objects.filter(est_actif=True).order_by('type', 'name')
    online_users = UserPresence.objects.filter(is_online=True).select_related('user')
    
    context = {
        'rooms': rooms,
        'online_users': online_users,
    }
    return render(request, 'chat/chat_home.html', context)


@login_required
def chat_room(request, room_slug):
    """Page d'un salon de chat"""
    room = get_object_or_404(ChatRoom, slug=room_slug, est_actif=True)
    room.membres.add(request.user)
    
    messages = room.messages.select_related('auteur').order_by('-date_envoi')[:50]
    messages = list(messages)[::-1]
    
    rooms = ChatRoom.objects.filter(est_actif=True).order_by('type', 'name')
    online_users = UserPresence.objects.filter(
        is_online=True, current_room=room
    ).select_related('user')
    
    context = {
        'room': room,
        'rooms': rooms,
        'messages': messages,
        'online_users': online_users,
    }
    return render(request, 'chat/chat_room.html', context)


@login_required
def chat_send_message(request):
    """API pour envoyer un message (fallback sans WebSocket)"""
    if request.method == 'POST':
        room_slug = request.POST.get('room_slug')
        content = request.POST.get('message', '').strip()
        
        if room_slug and content:
            room = get_object_or_404(ChatRoom, slug=room_slug)
            message = ChatMessage.objects.create(
                room=room,
                auteur=request.user,
                contenu=content,
                type_message='TEXT'
            )
            return JsonResponse({
                'success': True,
                'message_id': message.id,
                'timestamp': message.get_time_display(),
            })
    
    return JsonResponse({'success': False}, status=400)


@login_required
def chat_get_messages(request, room_slug):
    """API pour récupérer les nouveaux messages"""
    room = get_object_or_404(ChatRoom, slug=room_slug)
    last_id = request.GET.get('last_id', 0)
    
    messages = room.messages.filter(id__gt=last_id).select_related('auteur').order_by('date_envoi')
    
    data = [{
        'id': msg.id,
        'auteur': msg.auteur.username,
        'auteur_id': msg.auteur.id,
        'contenu': msg.contenu,
        'timestamp': msg.get_time_display(),
        'type': msg.type_message,
    } for msg in messages]
    
    return JsonResponse({'messages': data})


@login_required  
def send_system_notification(request):
    """Envoyer une notification système à tous les salons"""
    if request.method == 'POST' and request.user.is_staff:
        message = request.POST.get('message', '').strip()
        room_slug = request.POST.get('room_slug', 'general')
        
        if message:
            room = ChatRoom.objects.filter(slug=room_slug).first()
            if room:
                ChatMessage.objects.create(
                    room=room,
                    auteur=request.user,
                    contenu=message,
                    type_message='SYSTEM'
                )
                messages.success(request, "Notification envoyée !")
    
    return redirect('chat_home')